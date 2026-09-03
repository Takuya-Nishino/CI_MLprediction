# =============================================================================
# GitHub-ready analysis script
#
# Generated from the manuscript analysis notebook.
# Patient-level data are NOT included in this repository.
# Default paths are relative to the repository root:
#   data/Merge_20260901.xlsx
#   outputs/
#
# Before running, place an authorized local copy of the input dataset under
# data/ or edit DATA_PATH. Do not commit patient-level data or generated outputs.
# =============================================================================


# %% [markdown]
# # ③ Figures and tables — BMC MIDM / Methods-aligned / Standalone FULL
#
# Run after Notebooks 01 and 02. Creates publication tables, Additional-file tables, Figures 2–5 and supplementary figures, and exports Figure 1 cohort counts. It performs no model fitting. No external .py file is required.
#
# **Execution order:** 01 → 02 → 03. Each notebook contains the complete analysis function definitions and does not import `stroke_landmark_bmc_core_20260902.py` or `stroke_landmark_bmc_user_config_20260902.py`.
#
# **Windows/Jupyter stability fix:** SHAP is lazy-loaded and disabled by default in this reporting notebook. Main manuscript figures/tables are generated without SHAP. Set `RUN_SHAP=True` only if the optional SHAP figures are needed after the main stage runs successfully.
#
#
# **Submission-ready output update:** Main figures use a fixed color-blind–friendly palette and 170-mm publication width. The Day 7 operational matrix uses semantic four-stratum colors. Main tables are exported as editable Excel workbooks without cell shading/color fills.
#
# **FAST submission-ready reporting version:** optimized Notebook 03.
#
#
# **Final figure revision:** calibration legends contain model names only; Figure 5 is a two-panel Day 7 matrix comparing penalized logistic regression and LightGBM; all multi-panel labels use `(A)`, `(B)`, etc.; ROC/PR metrics are placed inside panels and DCA uses one common legend.

# %% 
# =============================================================================
# 0. Imports
# =============================================================================

from __future__ import annotations

import argparse
import os
import gc
import hashlib
import importlib.metadata as importlib_metadata
import json
import logging
import math
import platform
import re
import sys
import warnings
import zlib

# Windows/Jupyter stability: cap native numerical libraries before NumPy/LightGBM load.
# This changes only computational parallelism, not the statistical analysis.
for _var in (
    "OMP_NUM_THREADS",
    "MKL_NUM_THREADS",
    "OPENBLAS_NUM_THREADS",
    "NUMEXPR_NUM_THREADS",
    "VECLIB_MAXIMUM_THREADS",
    "BLIS_NUM_THREADS",
):
    os.environ[_var] = "1"
# Windows/Jupyter: avoid hard aborts when multiple OpenMP runtimes are loaded
# by NumPy/scikit-learn/LightGBM/SHAP. This notebook performs reporting only.
os.environ.setdefault("KMP_DUPLICATE_LIB_OK", "TRUE")
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any, Iterable, Optional
from threadpoolctl import threadpool_limits, threadpool_info

import joblib
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import sklearn
import statsmodels
import statsmodels.api as sm
from matplotlib.lines import Line2D
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sklearn.base import clone
from sklearn.compose import ColumnTransformer
from sklearn.impute import SimpleImputer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    auc,
    brier_score_loss,
    confusion_matrix,
    precision_recall_curve,
    roc_auc_score,
    roc_curve,
)
from sklearn.model_selection import (
    GridSearchCV,
    StratifiedKFold,
    cross_val_predict,
    cross_val_score,
)
from sklearn.pipeline import Pipeline
from sklearn.preprocessing import StandardScaler

warnings.filterwarnings("ignore", category=FutureWarning)
warnings.filterwarnings("ignore", message="X does not have valid feature names, but LGBMClassifier was fitted with feature names")
warnings.filterwarnings(
    "ignore",
    message="X does not have valid feature names, but LGBMClassifier was fitted with feature names",
    category=UserWarning,
)

# %% 
# =============================================================================
# 0B. Safe startup diagnostics
# =============================================================================
print("Notebook 03 safe-startup mode")
print("Python:", sys.version.split()[0])
print("Executable:", sys.executable)
print("Native thread limits:", {
    k: os.environ.get(k)
    for k in [
        "OMP_NUM_THREADS", "MKL_NUM_THREADS", "OPENBLAS_NUM_THREADS",
        "NUMEXPR_NUM_THREADS", "VECLIB_MAXIMUM_THREADS", "BLIS_NUM_THREADS"
    ]
})
print("KMP_DUPLICATE_LIB_OK:", os.environ.get("KMP_DUPLICATE_LIB_OK"))
print("SHAP is not imported at startup.")

# %% 
# =============================================================================
# 1. Configuration and prespecified variable definitions
# =============================================================================


@dataclass
class AnalysisConfig:
    """Analysis settings corresponding to the manuscript Methods."""

    data_path: Path = Path(
        r"data\Merge_20260901.xlsx"
    )
    output_dir: Path = Path(
        r"outputs"
    )
    mode: str = "FULL"  # FULL, LIGHT, or TEST

    # Cohort and dates
    id_col: str = "INDEX"
    patient_id_col: Optional[str] = "データ識別番号"
    admission_col: str = "入院日"
    discharge_col: str = "退院日"
    los_col: str = "在院日数"
    disposition_col: str = "転帰"
    development_start: str = "2018-04-01"
    temporal_cutoff: str = "2023-04-01"
    study_end_exclusive: str = "2025-04-01"
    enforce_first_hospitalization: bool = True
    known_duplicate_indices: tuple[Any, ...] = (4768,)

    # Reproducibility and computation
    seed: int = 20260902
    cv_folds: int = 5
    n_optuna_trials: int = 100
    n_bootstrap: int = 1000
    cv_n_jobs: int = 1
    lgbm_n_jobs: int = 2
    target_sensitivity: float = 0.80
    dca_threshold_min: float = 0.01
    dca_threshold_max: float = 0.70
    dca_threshold_step: float = 0.01

    # Prespecified model search space
    lr_c_grid: tuple[float, ...] = (
        0.01,
        0.03,
        0.1,
        0.3,
        1.0,
        3.0,
        10.0,
        30.0,
        100.0,
    )
    lgbm_learning_rate_min: float = 0.01
    lgbm_learning_rate_max: float = 0.20
    lgbm_estimators_min: int = 100
    lgbm_estimators_max: int = 800
    lgbm_max_depth_min: int = 2
    lgbm_max_depth_max: int = 6
    lgbm_min_child_samples_min: int = 20
    lgbm_min_child_samples_max: int = 100
    lgbm_subsample_min: float = 0.70
    lgbm_subsample_max: float = 1.00
    lgbm_colsample_min: float = 0.70
    lgbm_colsample_max: float = 1.00
    lgbm_regularization_min: float = 1e-4
    lgbm_regularization_max: float = 10.0

    # Analysis modules
    run_los28_day3: bool = False
    run_los28_day7: bool = True
    run_death_exclusion: bool = False
    run_shap: bool = True
    shap_max_patients: int = 2000

    # Figure output
    figure_dpi: int = 300
    figure_font: str = "Times New Roman"

    def apply_mode(self) -> "AnalysisConfig":
        mode = self.mode.upper()
        if mode == "FULL":
            return self
        if mode == "LIGHT":
            self.cv_folds = 3
            self.n_optuna_trials = 10
            self.n_bootstrap = 200
            self.lgbm_n_jobs = min(self.lgbm_n_jobs, 2)
            self.run_shap = False
            return self
        if mode == "TEST":
            self.cv_folds = 3
            self.n_optuna_trials = 1
            self.n_bootstrap = 12
            self.lgbm_n_jobs = 1
            self.run_shap = False
            self.lgbm_estimators_min = 15
            self.lgbm_estimators_max = 30
            self.lgbm_min_child_samples_min = 5
            self.lgbm_min_child_samples_max = 20
            return self
        raise ValueError("mode must be one of: FULL, LIGHT, TEST")


@dataclass(frozen=True)
class TaskSpec:
    key: str
    label: str
    landmark_day: int
    landmark_flag: str
    outcome_col: str
    outcome_label: str
    continuous_features: tuple[str, ...]
    binary_features: tuple[str, ...]
    exclusion_col: Optional[str] = None
    exclusion_value: Optional[int] = None
    analysis_type: str = "primary"

    @property
    def features(self) -> list[str]:
        return list(dict.fromkeys(self.continuous_features + self.binary_features))


# Baseline and landmark-specific features are retained exactly as in the
# established merged analysis dataset.
DAY3_CONTINUOUS: tuple[str, ...] = (
    "Age",
    "BMI",
    "NIHSS_total_24h後",
    "day3_Alb",
    "day3_BUN",
    "day3_eGFR",
    "day3_Hb",
    "day3_K",
    "day3_Na",
    "day3_WBC",
    "day3_CRP",
    "day3_SBP",
    "day3_HR",
    "day3_SU日数",
    "day3_B項目",
)

DAY3_BINARY: tuple[str, ...] = (
    "Male",
    "day3_t-pa",
    "day3_エダラボン",
    "day3_カテコラミン",
    "day3_抗生剤",
    "day3_抗てんかん剤",
    "day3_ハロペリドール_iv",
    "day3_脳血栓回収術",
    "day3_脳手術",
    "day3_人工呼吸",
    "day3_CHDF",
    "day3_頸動脈ステント留置術",
)

DAY7_CONTINUOUS: tuple[str, ...] = (
    "Age",
    "BMI",
    "NIHSS_total_24h後",
    "day7_Alb",
    "day7_BUN",
    "day7_eGFR",
    "day7_Hb",
    "day7_K",
    "day7_Na",
    "day7_WBC",
    "day7_CRP",
    "day7_SBP",
    "day7_HR",
    "day7_SU日数",
    "day7_B項目",
)

DAY7_BINARY: tuple[str, ...] = (
    "Male",
    "day7_t-pa",
    "day7_エダラボン",
    "day7_カテコラミン",
    "day7_抗生剤",
    "day7_抗てんかん剤",
    "day7_ハロペリドール_iv",
    "day7_脳血栓回収術",
    "day7_脳手術",
    "day7_人工呼吸",
    "day7_CHDF",
    "day7_頸動脈ステント留置術",
)

STRUCTURAL_ZERO_BINARY: tuple[str, ...] = tuple(
    c for c in DAY3_BINARY + DAY7_BINARY if c != "Male"
)
STRUCTURAL_ZERO_COUNT: tuple[str, ...] = ("day3_SU日数", "day7_SU日数")

PUBLICATION_LABELS: dict[str, str] = {
    "Age": "Age, years",
    "Male": "Male sex",
    "BMI": "Body mass index, kg/m²",
    "NIHSS_total_24h後": "NIHSS score at 24 h",
    "day3_Alb": "Albumin",
    "day3_BUN": "Blood urea nitrogen",
    "day3_eGFR": "eGFR",
    "day3_Hb": "Hemoglobin",
    "day3_K": "Potassium",
    "day3_Na": "Sodium",
    "day3_WBC": "White blood cell count",
    "day3_CRP": "C-reactive protein",
    "day3_SBP": "Systolic blood pressure",
    "day3_HR": "Heart rate",
    "day3_SU日数": "Stroke-unit days",
    "day3_B項目": "Nursing dependency score",
    "day3_t-pa": "Intravenous thrombolysis",
    "day3_エダラボン": "Edaravone",
    "day3_カテコラミン": "Catecholamine administration",
    "day3_抗生剤": "Systemic antibiotic therapy",
    "day3_抗てんかん剤": "Antiepileptic medication",
    "day3_ハロペリドール_iv": "Intravenous haloperidol",
    "day3_脳血栓回収術": "Endovascular thrombectomy",
    "day3_脳手術": "Neurosurgical procedure",
    "day3_人工呼吸": "Mechanical ventilation",
    "day3_CHDF": "Continuous hemodiafiltration",
    "day3_頸動脈ステント留置術": "Carotid artery stenting",
    "day7_Alb": "Albumin",
    "day7_BUN": "Blood urea nitrogen",
    "day7_eGFR": "eGFR",
    "day7_Hb": "Hemoglobin",
    "day7_K": "Potassium",
    "day7_Na": "Sodium",
    "day7_WBC": "White blood cell count",
    "day7_CRP": "C-reactive protein",
    "day7_SBP": "Systolic blood pressure",
    "day7_HR": "Heart rate",
    "day7_SU日数": "Stroke-unit days",
    "day7_B項目": "Nursing dependency score",
    "day7_t-pa": "Intravenous thrombolysis",
    "day7_エダラボン": "Edaravone",
    "day7_カテコラミン": "Catecholamine administration",
    "day7_抗生剤": "Systemic antibiotic therapy",
    "day7_抗てんかん剤": "Antiepileptic medication",
    "day7_ハロペリドール_iv": "Intravenous haloperidol",
    "day7_脳血栓回収術": "Endovascular thrombectomy",
    "day7_脳手術": "Neurosurgical procedure",
    "day7_人工呼吸": "Mechanical ventilation",
    "day7_CHDF": "Continuous hemodiafiltration",
    "day7_頸動脈ステント留置術": "Carotid artery stenting",
}


# Use a canonical module name for joblib bundles so that stage outputs remain
# readable whether a stage is run from Jupyter or through the command line.
_CANONICAL_MODULE_NAME = "stroke_landmark_bmc_core_20260902"
if __name__ == "__main__":
    sys.modules.setdefault(_CANONICAL_MODULE_NAME, sys.modules[__name__])
AnalysisConfig.__module__ = _CANONICAL_MODULE_NAME
TaskSpec.__module__ = _CANONICAL_MODULE_NAME

# %% 
# =============================================================================
# 2. General utilities, logging, and file output
# =============================================================================


def stable_seed(base_seed: int, text: str) -> int:
    """Derive a deterministic task-specific seed independent of Python hash."""
    return int((base_seed + zlib.crc32(text.encode("utf-8"))) % (2**31 - 1))


def setup_output_directories(config: AnalysisConfig) -> dict[str, Path]:
    root = Path(config.output_dir)
    paths = {
        "root": root,
        "models": root / "models",
        "predictions": root / "predictions",
        "tables": root / "tables",
        "figures": root / "figures",
        "diagnostics": root / "diagnostics",
        "bootstrap": root / "bootstrap",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def setup_logger(log_path: Path) -> logging.Logger:
    logger = logging.getLogger("stroke_landmark_bmc")
    logger.setLevel(logging.INFO)
    logger.handlers.clear()
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")

    stream = logging.StreamHandler(sys.stdout)
    stream.setFormatter(formatter)
    logger.addHandler(stream)

    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def save_dataframe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def save_figure(fig: plt.Figure, stem: Path, dpi: int) -> None:
    stem.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(stem.with_suffix(".pdf"), bbox_inches="tight")
    fig.savefig(
        stem.with_suffix(".tiff"),
        dpi=dpi,
        bbox_inches="tight",
        pil_kwargs={"compression": "tiff_lzw"},
    )
    plt.close(fig)


def json_safe(value: Any) -> Any:
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return None if not np.isfinite(value) else float(value)
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    return value


def _pkg_version(name: str) -> str:
    try:
        return importlib_metadata.version(name)
    except Exception:
        return "not available"


def package_versions() -> dict[str, str]:
    return {
        "python": platform.python_version(),
        "platform": platform.platform(),
        "numpy": np.__version__,
        "pandas": pd.__version__,
        "scikit-learn": sklearn.__version__,
        "lightgbm": _pkg_version("lightgbm"),
        "optuna": _pkg_version("optuna"),
        "statsmodels": statsmodels.__version__,
        "matplotlib": matplotlib.__version__,
        "shap": _pkg_version("shap"),
        "openpyxl": openpyxl.__version__,
        "joblib": joblib.__version__,
    }


def contains_japanese(value: Any) -> bool:
    if value is None:
        return False
    return bool(re.search(r"[\u3040-\u30ff\u3400-\u9fff]", str(value)))


def style_workbook(path: Path) -> None:
    """Apply BMC-compatible formatting without color or cell shading."""
    wb = openpyxl.load_workbook(path)
    align = Alignment(vertical="top", wrap_text=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                font_name = "Meiryo" if contains_japanese(cell.value) else "Times New Roman"
                cell.font = Font(name=font_name, bold=(row_idx == 1))
                cell.alignment = align

        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for column in ws.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell in column:
                    max_len = max(max_len, len("" if cell.value is None else str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    wb.save(path)


def write_excel_workbook(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    valid: dict[str, pd.DataFrame] = {}
    used: set[str] = set()
    for raw_name, df in sheets.items():
        if not isinstance(df, pd.DataFrame) or df.empty and len(df.columns) == 0:
            continue
        name = str(raw_name).replace("/", "_").replace("\\", "_")[:31] or "Sheet"
        base = name
        i = 1
        while name in used:
            suffix = f"_{i}"
            name = (base[: 31 - len(suffix)] + suffix)[:31]
            i += 1
        used.add(name)
        valid[name] = df

    if not valid:
        raise ValueError(f"No nonempty tables to write: {path}")

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in valid.items():
            df.to_excel(writer, sheet_name=name, index=False)
    style_workbook(path)

# %% 
# =============================================================================
# 3. Cohort construction, data validation, and outcome derivation
# =============================================================================


def coerce_male(series: pd.Series) -> pd.Series:
    """Convert a Male variable to 0/1 while rejecting ambiguous values."""
    s = series.copy()
    if pd.api.types.is_numeric_dtype(s):
        x = pd.to_numeric(s, errors="coerce")
        vals = set(x.dropna().unique().tolist())
        if vals.issubset({0, 1}):
            return x.astype(float)
        if vals.issubset({1, 2}) and 2 in vals:
            return x.map({1: 1.0, 2: 0.0})
        raise ValueError(f"Unexpected numeric Male values: {sorted(vals)}")

    txt = s.astype("string").str.strip().str.lower()
    mapping = {
        "男": 1.0,
        "男性": 1.0,
        "male": 1.0,
        "m": 1.0,
        "1": 1.0,
        "女": 0.0,
        "女性": 0.0,
        "female": 0.0,
        "f": 0.0,
        "0": 0.0,
        "2": 0.0,
    }
    out = txt.map(mapping)
    bad = s.notna() & out.isna()
    if bad.any():
        raise ValueError(
            "Unexpected Male labels: "
            + repr(sorted(s.loc[bad].astype(str).unique().tolist())[:20])
        )
    return out.astype(float)


def resolve_duplicate_indices(
    data: pd.DataFrame,
    config: AnalysisConfig,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Remove exact duplicates and explicitly handle only prespecified known IDs."""
    d = data.copy()
    id_col = config.id_col
    audit_rows: list[dict[str, Any]] = []

    exact_dup = d.duplicated(keep="first")
    if exact_dup.any():
        audit_rows.append(
            {
                "issue": "exact_duplicate_rows_removed",
                "n_rows": int(exact_dup.sum()),
                "details": "Rows identical across all columns",
            }
        )
        d = d.loc[~exact_dup].copy()

    duplicated_ids = d.loc[d[id_col].duplicated(keep=False), id_col].dropna().unique().tolist()
    unexpected = [x for x in duplicated_ids if x not in set(config.known_duplicate_indices)]
    if unexpected:
        raise ValueError(
            f"Conflicting duplicated {id_col} values were found and were not prespecified: "
            f"{unexpected[:20]}"
        )

    for value in duplicated_ids:
        n = int(d[id_col].eq(value).sum())
        audit_rows.append(
            {
                "issue": "known_duplicate_index_retained_first",
                "n_rows": n - 1,
                "details": f"{id_col}={value}; retained first row after source order",
            }
        )
        d = d.loc[~(d[id_col].eq(value) & d[id_col].duplicated(keep="first"))].copy()

    if d[id_col].duplicated().any():
        raise AssertionError(f"{id_col} remains duplicated after resolution")

    audit = pd.DataFrame(
        audit_rows,
        columns=["issue", "n_rows", "details"],
    )
    return d.reset_index(drop=True), audit


def retain_first_hospitalization(
    data: pd.DataFrame,
    config: AnalysisConfig,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    d = data.copy()
    if not config.enforce_first_hospitalization:
        return d, pd.DataFrame(
            [{"measure": "first_hospitalization_enforced", "value": False}]
        )

    patient_col = config.patient_id_col
    if patient_col is None:
        raise ValueError(
            "enforce_first_hospitalization=True requires patient_id_col to be specified"
        )
    if patient_col not in d.columns:
        warnings.warn(
            f"Patient identifier column '{patient_col}' was not found. "
            "The pipeline cannot independently verify retention of the first eligible "
            "hospitalization and therefore assumes that this restriction was already "
            "applied upstream.",
            RuntimeWarning,
        )
        return d, pd.DataFrame(
            [
                {
                    "measure": "first_hospitalization_enforced",
                    "value": False,
                },
                {
                    "measure": "reason",
                    "value": f"{patient_col} absent; assumed completed upstream",
                },
            ]
        )

    if d[patient_col].isna().any():
        raise ValueError(f"Missing values exist in patient identifier column: {patient_col}")

    before = len(d)
    patient_counts = d[patient_col].value_counts(dropna=False)
    repeated_patients = int(patient_counts.gt(1).sum())
    d = (
        d.sort_values(
            [patient_col, config.admission_col, config.id_col],
            kind="mergesort",
        )
        .drop_duplicates(subset=[patient_col], keep="first")
        .sort_values(config.id_col, kind="mergesort")
        .reset_index(drop=True)
    )
    audit = pd.DataFrame(
        [
            {"measure": "rows_before_first_hospitalization_filter", "value": before},
            {"measure": "patients_with_multiple_rows", "value": repeated_patients},
            {"measure": "rows_removed", "value": before - len(d)},
            {"measure": "rows_after_filter", "value": len(d)},
        ]
    )
    return d, audit


def load_and_prepare_data(
    config: AnalysisConfig,
    paths: dict[str, Path],
    logger: logging.Logger,
) -> tuple[pd.DataFrame, dict[str, pd.DataFrame]]:
    data_path = Path(config.data_path)
    if not data_path.exists():
        raise FileNotFoundError(f"Input file not found: {data_path}")

    logger.info("Loading input data: %s", data_path)
    if data_path.suffix.lower() in {".xlsx", ".xls"}:
        d = pd.read_excel(data_path)
    elif data_path.suffix.lower() == ".csv":
        d = pd.read_csv(data_path)
    elif data_path.suffix.lower() in {".parquet", ".pq"}:
        d = pd.read_parquet(data_path)
    else:
        raise ValueError("Supported input formats: xlsx, xls, csv, parquet")

    unnamed = [c for c in d.columns if str(c).startswith("Unnamed:")]
    if unnamed:
        d = d.drop(columns=unnamed)

    core_required = [
        config.id_col,
        config.admission_col,
        config.discharge_col,
        config.los_col,
        config.disposition_col,
    ]
    feature_required = sorted(set(DAY3_CONTINUOUS + DAY3_BINARY + DAY7_CONTINUOUS + DAY7_BINARY))
    required = core_required + feature_required
    # A patient identifier is used to enforce first-hospitalization retention when
    # available. It is not made a hard input requirement because some finalized
    # analytic files have already applied this restriction upstream.
    missing_cols = [c for c in required if c not in d.columns]
    if missing_cols:
        raise KeyError(f"Required columns are missing: {missing_cols}")

    # Dates and core numeric fields
    for col in [config.admission_col, config.discharge_col]:
        d[col] = pd.to_datetime(d[col], errors="coerce").dt.normalize()
    d[config.los_col] = pd.to_numeric(d[config.los_col], errors="coerce")

    if d[config.id_col].isna().any():
        raise ValueError(f"Missing values exist in {config.id_col}")
    if d[config.admission_col].isna().any():
        raise ValueError("Missing or invalid admission dates exist")
    if d[config.discharge_col].isna().any():
        raise ValueError("Missing or invalid discharge dates exist")
    if d[config.los_col].isna().any():
        raise ValueError("Missing or invalid length-of-stay values exist")
    if (d[config.discharge_col] < d[config.admission_col]).any():
        raise ValueError("At least one discharge date precedes the admission date")
    if (d[config.los_col] < 0).any():
        raise ValueError("Negative length-of-stay values exist")

    d, duplicate_audit = resolve_duplicate_indices(d, config)

    # Apply study period before retaining the first eligible hospitalization.
    dev_start = pd.Timestamp(config.development_start)
    cutoff = pd.Timestamp(config.temporal_cutoff)
    study_end = pd.Timestamp(config.study_end_exclusive)
    if not (dev_start < cutoff < study_end):
        raise ValueError("Study dates must satisfy development_start < temporal_cutoff < study_end")

    n_before_period = len(d)
    d = d.loc[
        d[config.admission_col].ge(dev_start)
        & d[config.admission_col].lt(study_end)
    ].copy()
    period_audit = pd.DataFrame(
        [
            {"measure": "rows_before_study_period_filter", "value": n_before_period},
            {"measure": "rows_after_study_period_filter", "value": len(d)},
        ]
    )

    d, first_admission_audit = retain_first_hospitalization(d, config)

    # Chronological split
    d["period"] = np.where(
        d[config.admission_col].lt(cutoff),
        "development",
        "temporal_evaluation",
    )
    if not set(d["period"].unique()).issubset({"development", "temporal_evaluation"}):
        raise AssertionError("Unexpected period labels")

    # Landmark dates and eligibility
    d["day3_landmark_date"] = d[config.admission_col] + pd.Timedelta(days=3)
    d["day7_landmark_date"] = d[config.admission_col] + pd.Timedelta(days=7)
    d["flag_day3_landmark"] = d[config.discharge_col].ge(d["day3_landmark_date"]).astype("int8")
    d["flag_day7_landmark"] = d[config.discharge_col].ge(d["day7_landmark_date"]).astype("int8")
    if (d["flag_day7_landmark"] > d["flag_day3_landmark"]).any():
        raise AssertionError("Day 7 landmark cohort is not nested within Day 3 cohort")

    # Outcome mapping: 0 or missing = home; 1 = transfer/non-home; 2 = in-hospital death.
    raw = d[config.disposition_col]
    numeric = pd.to_numeric(raw, errors="coerce")
    invalid_non_numeric = raw.notna() & numeric.isna()
    if invalid_non_numeric.any():
        raise ValueError(
            "Non-numeric values were found in 転帰 despite the prespecified coding: "
            + repr(sorted(raw.loc[invalid_non_numeric].astype(str).unique().tolist())[:20])
        )
    observed_codes = set(numeric.dropna().astype(int).unique().tolist())
    if not observed_codes.issubset({0, 1, 2}):
        raise ValueError(f"Unexpected 転帰 codes: {sorted(observed_codes)}")
    d["disposition_code"] = numeric
    d["is_inhospital_death"] = numeric.eq(2).astype("int8")
    d["Outcome_nonhome"] = numeric.isin([1, 2]).astype("int8")
    d["Outcome_transfer_no_death"] = numeric.eq(1).astype("int8")
    d["LOS_21"] = d[config.los_col].gt(21).astype("int8")
    d["LOS_28"] = d[config.los_col].gt(28).astype("int8")

    # Feature type conversion and structural-zero handling.
    d["Male"] = coerce_male(d["Male"])

    for col in STRUCTURAL_ZERO_BINARY:
        x = pd.to_numeric(d[col], errors="coerce")
        if (x.dropna() < 0).any():
            raise ValueError(f"Negative treatment/procedure values in {col}")
        d[col] = np.where(x.isna(), 0, (x > 0).astype(int)).astype("int8")

    for col in STRUCTURAL_ZERO_COUNT:
        x = pd.to_numeric(d[col], errors="coerce")
        if (x.dropna() < 0).any():
            raise ValueError(f"Negative stroke-unit day values in {col}")
        d[col] = x.fillna(0).astype(float)

    all_cont = sorted(set(DAY3_CONTINUOUS + DAY7_CONTINUOUS))
    for col in all_cont:
        if col in STRUCTURAL_ZERO_COUNT:
            continue
        d[col] = pd.to_numeric(d[col], errors="coerce").replace([np.inf, -np.inf], np.nan)

    # Binary input QA. Male may remain missing and will be imputed within folds.
    for col in sorted(set(DAY3_BINARY + DAY7_BINARY)):
        values = set(pd.to_numeric(d[col], errors="coerce").dropna().unique().tolist())
        if not values.issubset({0, 1, 0.0, 1.0}):
            raise ValueError(f"Binary predictor {col} contains unexpected values: {sorted(values)[:20]}")

    # Landmark-cumulative variables must not decrease between Day 3 and Day 7
    # among patients who reached Day 7. This verifies the upstream wide-variable
    # construction assumed by the Methods.
    cumulative_audit_rows: list[dict[str, Any]] = []
    day7_mask = d["flag_day7_landmark"].eq(1)
    cumulative_pairs = [
        (f"day3_{suffix}", f"day7_{suffix}")
        for suffix in [
            "t-pa", "エダラボン", "カテコラミン", "抗生剤",
            "抗てんかん剤", "ハロペリドール_iv", "脳血栓回収術",
            "脳手術", "人工呼吸", "CHDF", "頸動脈ステント留置術",
        ]
    ] + [("day3_SU日数", "day7_SU日数")]
    for day3_col, day7_col in cumulative_pairs:
        violation = day7_mask & d[day7_col].lt(d[day3_col])
        n_violation = int(violation.sum())
        cumulative_audit_rows.append(
            {
                "Day3_variable": day3_col,
                "Day7_variable": day7_col,
                "N_Day7_eligible": int(day7_mask.sum()),
                "N_decreasing": n_violation,
            }
        )
        if n_violation:
            examples = d.loc[violation, [config.id_col, day3_col, day7_col]].head(10)
            raise ValueError(
                f"Cumulative landmark variable decreases from Day 3 to Day 7: "
                f"{day3_col} -> {day7_col}; n={n_violation}. Examples:\n{examples}"
            )
    cumulative_landmark_audit = pd.DataFrame(cumulative_audit_rows)

    # LOS/date discrepancy is audited but the supplied LOS field remains the outcome source.
    d["calendar_date_difference"] = (
        d[config.discharge_col] - d[config.admission_col]
    ).dt.days
    d["los_minus_date_difference"] = d[config.los_col] - d["calendar_date_difference"]
    los_date_audit = (
        d["los_minus_date_difference"]
        .value_counts(dropna=False)
        .rename_axis("LOS minus calendar date difference")
        .reset_index(name="n")
        .sort_values("LOS minus calendar date difference")
    )

    flow = make_flow_table(d, config)
    outcome_audit = make_outcome_audit(d, config)
    disposition_audit = (
        d[config.disposition_col]
        .value_counts(dropna=False)
        .rename_axis("raw_disposition_code")
        .reset_index(name="n")
    )

    logger.info("Eligible study-period hospitalizations after patient-level deduplication: %d", len(d))
    logger.info("Day 3 landmark cohort: %d", int(d["flag_day3_landmark"].sum()))
    logger.info("Day 7 landmark cohort: %d", int(d["flag_day7_landmark"].sum()))

    audits = {
        "duplicate_audit": duplicate_audit,
        "period_audit": period_audit,
        "first_hospitalization_audit": first_admission_audit,
        "los_date_audit": los_date_audit,
        "flow": flow,
        "outcome_audit": outcome_audit,
        "disposition_audit": disposition_audit,
        "cumulative_landmark_audit": cumulative_landmark_audit,
    }
    for name, table in audits.items():
        save_dataframe(table, paths["diagnostics"] / f"{name}.csv")

    return d.reset_index(drop=True), audits


def make_flow_table(data: pd.DataFrame, config: AnalysisConfig) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    n_total = len(data)
    n_d3 = int(data["flag_day3_landmark"].sum())
    n_d7 = int(data["flag_day7_landmark"].sum())
    rows.extend(
        [
            {"Stage": "Eligible acute stroke admissions", "N": n_total},
            {"Stage": "Discharged before Day 3", "N": n_total - n_d3},
            {"Stage": "Reached Day 3 landmark", "N": n_d3},
            {"Stage": "Discharged after Day 3 but before Day 7", "N": n_d3 - n_d7},
            {"Stage": "Reached Day 7 landmark", "N": n_d7},
        ]
    )
    for day, flag in [(3, "flag_day3_landmark"), (7, "flag_day7_landmark")]:
        for period in ["development", "temporal_evaluation"]:
            rows.append(
                {
                    "Stage": f"Day {day}: {period}",
                    "N": int(data[flag].eq(1).mul(data["period"].eq(period)).sum()),
                }
            )
    return pd.DataFrame(rows)


def make_outcome_audit(data: pd.DataFrame, config: AnalysisConfig) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for day, flag in [(3, "flag_day3_landmark"), (7, "flag_day7_landmark")]:
        for period in ["development", "temporal_evaluation"]:
            sub = data.loc[data[flag].eq(1) & data["period"].eq(period)]
            rows.append(
                {
                    "Landmark": f"Day {day}",
                    "Period": period,
                    "N": len(sub),
                    "Nonhome_events": int(sub["Outcome_nonhome"].sum()),
                    "Nonhome_rate": float(sub["Outcome_nonhome"].mean()),
                    "Transfer_events_excluding_death": int(sub["Outcome_transfer_no_death"].sum()),
                    "Inhospital_deaths": int(sub["is_inhospital_death"].sum()),
                    "LOS21_events": int(sub["LOS_21"].sum()),
                    "LOS21_rate": float(sub["LOS_21"].mean()),
                    "LOS28_events": int(sub["LOS_28"].sum()),
                    "LOS28_rate": float(sub["LOS_28"].mean()),
                    "Median_LOS": float(sub[config.los_col].median()),
                    "Q1_LOS": float(sub[config.los_col].quantile(0.25)),
                    "Q3_LOS": float(sub[config.los_col].quantile(0.75)),
                }
            )
    return pd.DataFrame(rows)

# %% 
# =============================================================================
# 4. Task specifications and predictor dictionary
# =============================================================================


def build_task_specs(config: AnalysisConfig) -> tuple[list[TaskSpec], list[TaskSpec]]:
    primary = [
        TaskSpec(
            key="day3_nonhome",
            label="Day 3: no direct home discharge",
            landmark_day=3,
            landmark_flag="flag_day3_landmark",
            outcome_col="Outcome_nonhome",
            outcome_label="No direct home discharge",
            continuous_features=DAY3_CONTINUOUS,
            binary_features=DAY3_BINARY,
        ),
        TaskSpec(
            key="day3_los21",
            label="Day 3: prolonged hospitalization >21 days",
            landmark_day=3,
            landmark_flag="flag_day3_landmark",
            outcome_col="LOS_21",
            outcome_label="Hospital length of stay >21 days",
            continuous_features=DAY3_CONTINUOUS,
            binary_features=DAY3_BINARY,
        ),
        TaskSpec(
            key="day7_nonhome",
            label="Day 7: no direct home discharge",
            landmark_day=7,
            landmark_flag="flag_day7_landmark",
            outcome_col="Outcome_nonhome",
            outcome_label="No direct home discharge",
            continuous_features=DAY7_CONTINUOUS,
            binary_features=DAY7_BINARY,
        ),
        TaskSpec(
            key="day7_los21",
            label="Day 7: prolonged hospitalization >21 days",
            landmark_day=7,
            landmark_flag="flag_day7_landmark",
            outcome_col="LOS_21",
            outcome_label="Hospital length of stay >21 days",
            continuous_features=DAY7_CONTINUOUS,
            binary_features=DAY7_BINARY,
        ),
    ]

    sensitivity: list[TaskSpec] = []
    if config.run_los28_day3:
        sensitivity.append(
            TaskSpec(
                key="day3_los28",
                label="Sensitivity: Day 3 prolonged hospitalization >28 days",
                landmark_day=3,
                landmark_flag="flag_day3_landmark",
                outcome_col="LOS_28",
                outcome_label="Hospital length of stay >28 days",
                continuous_features=DAY3_CONTINUOUS,
                binary_features=DAY3_BINARY,
                analysis_type="sensitivity_los28",
            )
        )
    if config.run_los28_day7:
        sensitivity.append(
            TaskSpec(
                key="day7_los28",
                label="Sensitivity: Day 7 prolonged hospitalization >28 days",
                landmark_day=7,
                landmark_flag="flag_day7_landmark",
                outcome_col="LOS_28",
                outcome_label="Hospital length of stay >28 days",
                continuous_features=DAY7_CONTINUOUS,
                binary_features=DAY7_BINARY,
                analysis_type="sensitivity_los28",
            )
        )
    if config.run_death_exclusion:
        sensitivity.extend(
            [
                TaskSpec(
                    key="day3_nonhome_excluding_death",
                    label="Sensitivity: Day 3 transfer versus home, deaths excluded",
                    landmark_day=3,
                    landmark_flag="flag_day3_landmark",
                    outcome_col="Outcome_transfer_no_death",
                    outcome_label="Transfer versus direct home discharge",
                    continuous_features=DAY3_CONTINUOUS,
                    binary_features=DAY3_BINARY,
                    exclusion_col="is_inhospital_death",
                    exclusion_value=1,
                    analysis_type="sensitivity_death_exclusion",
                ),
                TaskSpec(
                    key="day7_nonhome_excluding_death",
                    label="Sensitivity: Day 7 transfer versus home, deaths excluded",
                    landmark_day=7,
                    landmark_flag="flag_day7_landmark",
                    outcome_col="Outcome_transfer_no_death",
                    outcome_label="Transfer versus direct home discharge",
                    continuous_features=DAY7_CONTINUOUS,
                    binary_features=DAY7_BINARY,
                    exclusion_col="is_inhospital_death",
                    exclusion_value=1,
                    analysis_type="sensitivity_death_exclusion",
                ),
            ]
        )
    return primary, sensitivity


def make_predictor_dictionary() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    base_defs = {
        "Age": ("Patient characteristics", "Age at admission", "Continuous"),
        "Male": ("Patient characteristics", "Male sex", "Binary"),
        "BMI": ("Patient characteristics", "Body mass index", "Continuous"),
        "NIHSS_total_24h後": (
            "Neurological severity",
            "NIHSS total score assessed approximately 24 h after admission",
            "Ordinal/continuous",
        ),
    }
    for raw, (domain, definition, dtype) in base_defs.items():
        rows.append(
            {
                "Raw_variable": raw,
                "Publication_label": PUBLICATION_LABELS[raw],
                "Domain": domain,
                "Definition": definition,
                "Type": dtype,
                "Day3": "Yes",
                "Day7": "Yes",
                "Missing_value_handling": "Median" if raw != "Male" else "Most frequent",
            }
        )

    domain_maps = [
        (
            ["Alb", "BUN", "eGFR", "Hb", "K", "Na", "WBC", "CRP"],
            "Laboratory findings",
            "Most recent measurement available by the landmark",
            "Continuous",
            "Median",
        ),
        (
            ["SBP", "HR"],
            "Vital signs",
            "Most recent measurement available by the landmark",
            "Continuous",
            "Median",
        ),
        (
            ["SU日数"],
            "Stroke-unit utilization",
            "Cumulative calendar days in the stroke-unit environment by the landmark",
            "Count",
            "Structural zero if no stay",
        ),
        (
            ["B項目"],
            "Nursing dependency",
            "Most recent total B-item score available by the landmark",
            "Ordinal/continuous",
            "Median",
        ),
    ]
    for suffixes, domain, definition, dtype, missing in domain_maps:
        for suffix in suffixes:
            for day in [3, 7]:
                raw = f"day{day}_{suffix}"
                rows.append(
                    {
                        "Raw_variable": raw,
                        "Publication_label": PUBLICATION_LABELS[raw],
                        "Domain": domain,
                        "Definition": definition,
                        "Type": dtype,
                        "Day3": "Yes" if day == 3 else "No",
                        "Day7": "Yes" if day == 7 else "No",
                        "Missing_value_handling": missing,
                    }
                )

    medication_suffixes = [
        "t-pa",
        "エダラボン",
        "カテコラミン",
        "抗生剤",
        "抗てんかん剤",
        "ハロペリドール_iv",
    ]
    procedure_suffixes = [
        "脳血栓回収術",
        "脳手術",
        "人工呼吸",
        "CHDF",
        "頸動脈ステント留置術",
    ]
    for suffixes, domain in [
        (medication_suffixes, "Acute medication"),
        (procedure_suffixes, "Procedure"),
    ]:
        for suffix in suffixes:
            for day in [3, 7]:
                raw = f"day{day}_{suffix}"
                rows.append(
                    {
                        "Raw_variable": raw,
                        "Publication_label": PUBLICATION_LABELS[raw],
                        "Domain": domain,
                        "Definition": "At least one qualifying record from admission through the landmark",
                        "Type": "Binary",
                        "Day3": "Yes" if day == 3 else "No",
                        "Day7": "Yes" if day == 7 else "No",
                        "Missing_value_handling": "Structural zero if no record",
                    }
                )
    return pd.DataFrame(rows)

# %% 
# =============================================================================
# 5. Performance metrics, calibration, thresholds, bootstrap, and DCA
# =============================================================================


def pr_auc_score(y_true: Iterable[int], y_prob: Iterable[float]) -> float:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    precision, recall, _ = precision_recall_curve(y, p)
    order = np.argsort(recall)
    return float(auc(recall[order], precision[order]))


def calibration_intercept_slope(
    y_true: Iterable[int],
    y_prob: Iterable[float],
) -> tuple[float, float]:
    y = np.asarray(y_true, dtype=float)
    p = np.clip(np.asarray(y_prob, dtype=float), 1e-6, 1 - 1e-6)
    lp = np.log(p / (1 - p))

    intercept = np.nan
    slope = np.nan
    try:
        fit_int = sm.GLM(
            y,
            np.ones((len(y), 1)),
            family=sm.families.Binomial(),
            offset=lp,
        ).fit(disp=0)
        intercept = float(fit_int.params[0])
    except Exception:
        pass

    try:
        fit_slope = sm.GLM(
            y,
            sm.add_constant(lp, has_constant="add"),
            family=sm.families.Binomial(),
        ).fit(disp=0)
        slope = float(fit_slope.params[1])
    except Exception:
        pass
    return intercept, slope


def threshold_metrics(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    threshold: float,
) -> dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    pred = (p >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y, pred, labels=[0, 1]).ravel()
    sensitivity = tp / (tp + fn) if tp + fn else np.nan
    specificity = tn / (tn + fp) if tn + fp else np.nan
    ppv = tp / (tp + fp) if tp + fp else np.nan
    npv = tn / (tn + fn) if tn + fn else np.nan
    return {
        "Sensitivity": float(sensitivity),
        "Specificity": float(specificity),
        "PPV": float(ppv),
        "NPV": float(npv),
        "TP": float(tp),
        "FP": float(fp),
        "TN": float(tn),
        "FN": float(fn),
    }


def calculate_metrics(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    threshold: Optional[float] = None,
) -> dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    if len(y) != len(p):
        raise ValueError("y_true and y_prob have different lengths")
    if np.unique(y).size < 2:
        raise ValueError("Performance metrics require both outcome classes")
    if not np.isfinite(p).all() or ((p < 0) | (p > 1)).any():
        raise ValueError("Predicted probabilities must be finite and within [0,1]")

    intercept, slope = calibration_intercept_slope(y, p)
    out: dict[str, float] = {
        "AUROC": float(roc_auc_score(y, p)),
        "PR_AUC": pr_auc_score(y, p),
        "Brier": float(brier_score_loss(y, p)),
        "Calibration_intercept": intercept,
        "Calibration_slope": slope,
    }
    if threshold is not None:
        out.update(threshold_metrics(y, p, threshold))
    return out


def choose_threshold_at_target_sensitivity(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    target: float,
) -> dict[str, float]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    fpr, tpr, thresholds = roc_curve(y, p)
    valid = np.isfinite(thresholds)
    fpr, tpr, thresholds = fpr[valid], tpr[valid], thresholds[valid]
    if len(thresholds) == 0:
        raise RuntimeError("No finite ROC threshold was available")
    distance = np.abs(tpr - target)
    min_distance = float(distance.min())
    candidates = np.flatnonzero(np.isclose(distance, min_distance))
    # Tie-breaker: higher specificity, then higher threshold.
    specificity = 1 - fpr
    best_specificity = specificity[candidates].max()
    candidates = candidates[np.isclose(specificity[candidates], best_specificity)]
    idx = candidates[np.argmax(thresholds[candidates])]
    return {
        "threshold": float(thresholds[idx]),
        "development_oof_sensitivity": float(tpr[idx]),
        "development_oof_specificity": float(specificity[idx]),
        "absolute_distance_from_target": float(distance[idx]),
    }


def bootstrap_metrics(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    threshold: float,
    n_bootstrap: int,
    seed: int,
) -> tuple[pd.DataFrame, dict[str, tuple[float, float]]]:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    n = len(y)
    rng = np.random.default_rng(seed)
    rows: list[dict[str, float]] = []
    max_attempts = n_bootstrap * 10
    attempts = 0

    while len(rows) < n_bootstrap and attempts < max_attempts:
        attempts += 1
        idx = rng.integers(0, n, size=n)
        yb = y[idx]
        if np.unique(yb).size < 2:
            continue
        try:
            rows.append(calculate_metrics(yb, p[idx], threshold=threshold))
        except Exception:
            continue

    minimum_required = min(n_bootstrap, max(20, int(math.ceil(n_bootstrap * 0.80))))
    if len(rows) < minimum_required:
        raise RuntimeError(
            f"Only {len(rows)} valid bootstrap resamples were obtained; "
            f"required at least {minimum_required}."
        )

    boot = pd.DataFrame(rows)
    ci: dict[str, tuple[float, float]] = {}
    for col in boot.columns:
        values = boot[col].replace([np.inf, -np.inf], np.nan).dropna().to_numpy()
        if len(values) < max(10, int(0.50 * len(boot))):
            ci[col] = (np.nan, np.nan)
        else:
            ci[col] = (
                float(np.percentile(values, 2.5)),
                float(np.percentile(values, 97.5)),
            )
    return boot, ci


def calibration_curve_quantile(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    n_bins: int = 10,
) -> pd.DataFrame:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    if len(y) == 0:
        return pd.DataFrame(columns=["bin", "n", "mean_predicted", "observed"])
    q = min(n_bins, len(y))
    ranks = pd.Series(p).rank(method="first")
    bins = pd.qcut(ranks, q=q, labels=False, duplicates="drop")
    df = pd.DataFrame({"y": y, "p": p, "bin": bins})
    out = (
        df.groupby("bin", observed=True)
        .agg(n=("y", "size"), mean_predicted=("p", "mean"), observed=("y", "mean"))
        .reset_index()
    )
    return out


def decision_curve(
    y_true: Iterable[int],
    y_prob: Iterable[float],
    thresholds: np.ndarray,
) -> pd.DataFrame:
    y = np.asarray(y_true, dtype=int)
    p = np.asarray(y_prob, dtype=float)
    n = len(y)
    prevalence = float(y.mean())
    rows: list[dict[str, float]] = []
    for pt in thresholds:
        pred = p >= pt
        tp = int(np.sum(pred & (y == 1)))
        fp = int(np.sum(pred & (y == 0)))
        odds = pt / (1 - pt)
        rows.append(
            {
                "threshold": float(pt),
                "model": float(tp / n - fp / n * odds),
                "treat_all": float(prevalence - (1 - prevalence) * odds),
                "treat_none": 0.0,
            }
        )
    return pd.DataFrame(rows)

# %% 
# =============================================================================
# 6. Model pipelines, tuning, and task fitting
# =============================================================================


def make_preprocessor(
    continuous_features: list[str],
    binary_features: list[str],
    standardize_continuous: bool,
) -> ColumnTransformer:
    continuous_steps: list[tuple[str, Any]] = [
        ("imputer", SimpleImputer(strategy="median", keep_empty_features=True)),
    ]
    if standardize_continuous:
        continuous_steps.append(("scaler", StandardScaler()))

    binary_steps: list[tuple[str, Any]] = [
        ("imputer", SimpleImputer(strategy="most_frequent", keep_empty_features=True)),
    ]
    return ColumnTransformer(
        transformers=[
            ("continuous", Pipeline(continuous_steps), continuous_features),
            ("binary", Pipeline(binary_steps), binary_features),
        ],
        remainder="drop",
        verbose_feature_names_out=True,
    )


def make_lr_pipeline(
    continuous_features: list[str],
    binary_features: list[str],
    c_value: float,
    seed: int,
) -> Pipeline:
    return Pipeline(
        steps=[
            (
                "preprocess",
                make_preprocessor(
                    continuous_features,
                    binary_features,
                    standardize_continuous=True,
                ),
            ),
            (
                "model",
                LogisticRegression(
                    penalty="l2",
                    C=c_value,
                    solver="lbfgs",
                    max_iter=5000,
                    random_state=seed,
                ),
            ),
        ]
    )


def make_lgbm_pipeline(
    continuous_features: list[str],
    binary_features: list[str],
    params: dict[str, Any],
    seed: int,
    n_jobs: int,
) -> Pipeline:
    model_params = {
        "objective": "binary",
        "verbosity": -1,
        "random_state": seed,
        "n_jobs": n_jobs,
        "subsample_freq": 1,
        "deterministic": True,
        "force_col_wise": True,
        "feature_fraction_seed": seed,
        "bagging_seed": seed,
        "data_random_seed": seed,
        **params,
    }
    return Pipeline(
        steps=[
            (
                "preprocess",
                make_preprocessor(
                    continuous_features,
                    binary_features,
                    standardize_continuous=False,
                ),
            ),
            ("model", LGBMClassifier(**model_params)),
        ]
    )


def check_cv_feasibility(y: pd.Series, n_splits: int, task_label: str) -> None:
    counts = y.value_counts().sort_index()
    if len(counts) != 2:
        raise ValueError(f"{task_label}: outcome does not contain exactly two classes")
    if int(counts.min()) < n_splits:
        raise ValueError(
            f"{task_label}: smallest class has {int(counts.min())} observations, "
            f"which is fewer than cv_folds={n_splits}."
        )


def tune_logistic_regression(
    x: pd.DataFrame,
    y: pd.Series,
    continuous_features: list[str],
    binary_features: list[str],
    config: AnalysisConfig,
    task_seed: int,
) -> tuple[Pipeline, dict[str, Any], np.ndarray, pd.DataFrame]:
    cv = StratifiedKFold(
        n_splits=config.cv_folds,
        shuffle=True,
        random_state=task_seed,
    )
    base = make_lr_pipeline(
        continuous_features,
        binary_features,
        c_value=1.0,
        seed=task_seed,
    )
    search = GridSearchCV(
        estimator=base,
        param_grid={"model__C": list(config.lr_c_grid)},
        scoring="neg_log_loss",
        cv=cv,
        refit=True,
        n_jobs=config.cv_n_jobs,
        return_train_score=False,
        error_score="raise",
    )
    search.fit(x, y)
    best_c = float(search.best_params_["model__C"])
    final_pipeline = make_lr_pipeline(
        continuous_features,
        binary_features,
        c_value=best_c,
        seed=task_seed,
    )
    oof = cross_val_predict(
        clone(final_pipeline),
        x,
        y,
        cv=cv,
        method="predict_proba",
        n_jobs=config.cv_n_jobs,
    )[:, 1]
    final_pipeline.fit(x, y)
    cv_results = pd.DataFrame(search.cv_results_)
    keep = [
        "param_model__C",
        "mean_test_score",
        "std_test_score",
        "rank_test_score",
    ]
    cv_results = cv_results[keep].rename(
        columns={
            "param_model__C": "C",
            "mean_test_score": "mean_negative_log_loss",
            "std_test_score": "sd_negative_log_loss",
            "rank_test_score": "rank",
        }
    )
    cv_results["mean_cv_log_loss"] = -cv_results["mean_negative_log_loss"]
    params = {
        "C": best_c,
        "mean_cv_log_loss": float(-search.best_score_),
        "cv_folds": config.cv_folds,
    }
    return final_pipeline, params, oof, cv_results


def tune_lightgbm(
    x: pd.DataFrame,
    y: pd.Series,
    continuous_features: list[str],
    binary_features: list[str],
    config: AnalysisConfig,
    task_seed: int,
) -> tuple[Pipeline, dict[str, Any], np.ndarray, pd.DataFrame]:
    cv = StratifiedKFold(
        n_splits=config.cv_folds,
        shuffle=True,
        random_state=task_seed,
    )

    def objective(trial: optuna.Trial) -> float:
        max_depth = trial.suggest_int(
            "max_depth",
            config.lgbm_max_depth_min,
            config.lgbm_max_depth_max,
        )
        max_leaves = min(63, 2**max_depth)
        params = {
            "learning_rate": trial.suggest_float(
                "learning_rate",
                config.lgbm_learning_rate_min,
                config.lgbm_learning_rate_max,
                log=True,
            ),
            "n_estimators": trial.suggest_int(
                "n_estimators",
                config.lgbm_estimators_min,
                config.lgbm_estimators_max,
            ),
            "max_depth": max_depth,
            "num_leaves": trial.suggest_int("num_leaves", 3, max_leaves),
            "min_child_samples": trial.suggest_int(
                "min_child_samples",
                config.lgbm_min_child_samples_min,
                config.lgbm_min_child_samples_max,
            ),
            "subsample": trial.suggest_float(
                "subsample",
                config.lgbm_subsample_min,
                config.lgbm_subsample_max,
            ),
            "colsample_bytree": trial.suggest_float(
                "colsample_bytree",
                config.lgbm_colsample_min,
                config.lgbm_colsample_max,
            ),
            "reg_alpha": trial.suggest_float(
                "reg_alpha",
                config.lgbm_regularization_min,
                config.lgbm_regularization_max,
                log=True,
            ),
            "reg_lambda": trial.suggest_float(
                "reg_lambda",
                config.lgbm_regularization_min,
                config.lgbm_regularization_max,
                log=True,
            ),
        }
        model = make_lgbm_pipeline(
            continuous_features,
            binary_features,
            params=params,
            seed=task_seed,
            n_jobs=config.lgbm_n_jobs,
        )
        scores = cross_val_score(
            model,
            x,
            y,
            scoring="neg_log_loss",
            cv=cv,
            n_jobs=config.cv_n_jobs,
            error_score="raise",
        )
        return float(-scores.mean())

    study = optuna.create_study(
        direction="minimize",
        sampler=optuna.samplers.TPESampler(seed=task_seed),
    )
    study.optimize(
        objective,
        n_trials=config.n_optuna_trials,
        show_progress_bar=False,
    )
    best_model_params = dict(study.best_params)
    # Defensive QA for the previous max_depth/num_leaves inconsistency.
    if best_model_params["num_leaves"] > 2 ** best_model_params["max_depth"]:
        raise AssertionError("Incoherent LightGBM max_depth and num_leaves configuration")

    final_pipeline = make_lgbm_pipeline(
        continuous_features,
        binary_features,
        params=best_model_params,
        seed=task_seed,
        n_jobs=config.lgbm_n_jobs,
    )
    oof = cross_val_predict(
        clone(final_pipeline),
        x,
        y,
        cv=cv,
        method="predict_proba",
        n_jobs=config.cv_n_jobs,
    )[:, 1]
    final_pipeline.fit(x, y)

    trials = study.trials_dataframe(
        attrs=("number", "value", "params", "state")
    )
    params = {
        **best_model_params,
        "mean_cv_log_loss": float(study.best_value),
        "cv_folds": config.cv_folds,
        "n_trials": config.n_optuna_trials,
        "subsample_freq": 1,
    }
    return final_pipeline, params, oof, trials


def fit_prediction_task(
    data: pd.DataFrame,
    spec: TaskSpec,
    config: AnalysisConfig,
    paths: dict[str, Path],
    logger: logging.Logger,
) -> dict[str, Any]:
    task_seed = stable_seed(config.seed, spec.key)
    mask = data[spec.landmark_flag].eq(1) & data[spec.outcome_col].notna()
    if spec.exclusion_col is not None:
        mask &= ~data[spec.exclusion_col].eq(spec.exclusion_value)
    task_data = data.loc[mask].copy()

    dev = task_data.loc[task_data["period"].eq("development")].copy()
    temp = task_data.loc[task_data["period"].eq("temporal_evaluation")].copy()
    if dev.empty or temp.empty:
        raise ValueError(f"{spec.label}: empty development or temporal cohort")
    if dev[config.admission_col].max() >= temp[config.admission_col].min():
        raise AssertionError(f"{spec.label}: chronological periods overlap")
    if set(dev[config.id_col]).intersection(set(temp[config.id_col])):
        raise AssertionError(f"{spec.label}: IDs overlap across periods")

    x_dev = dev[spec.features].copy()
    y_dev = dev[spec.outcome_col].astype(int).copy()
    x_temp = temp[spec.features].copy()
    y_temp = temp[spec.outcome_col].astype(int).copy()

    # Median or most-frequent imputation is undefined when a predictor is
    # completely missing in the development cohort. Stop explicitly rather
    # than allowing a library-dependent constant to be inserted silently.
    all_missing_development = [
        col for col in spec.features if x_dev[col].isna().all()
    ]
    if all_missing_development:
        raise ValueError(
            f"{spec.label}: predictors entirely missing in development data: "
            f"{all_missing_development}"
        )

    check_cv_feasibility(y_dev, config.cv_folds, spec.label)
    if y_temp.nunique() != 2:
        raise ValueError(f"{spec.label}: temporal evaluation has fewer than two outcome classes")

    logger.info(
        "%s | development n=%d events=%d | temporal n=%d events=%d",
        spec.label,
        len(dev),
        int(y_dev.sum()),
        len(temp),
        int(y_temp.sum()),
    )

    lr_model, lr_params, lr_oof, lr_cv = tune_logistic_regression(
        x_dev,
        y_dev,
        list(spec.continuous_features),
        list(spec.binary_features),
        config,
        task_seed,
    )
    lgb_model, lgb_params, lgb_oof, lgb_trials = tune_lightgbm(
        x_dev,
        y_dev,
        list(spec.continuous_features),
        list(spec.binary_features),
        config,
        task_seed + 17,
    )

    models: dict[str, dict[str, Any]] = {}
    for model_key, model_label, model, params, oof, tuning in [
        ("lr", "Penalized logistic regression", lr_model, lr_params, lr_oof, lr_cv),
        ("lgbm", "LightGBM", lgb_model, lgb_params, lgb_oof, lgb_trials),
    ]:
        temporal_prob = model.predict_proba(x_temp)[:, 1]
        threshold_info = choose_threshold_at_target_sensitivity(
            y_dev,
            oof,
            target=config.target_sensitivity,
        )
        threshold = threshold_info["threshold"]
        point = calculate_metrics(y_temp, temporal_prob, threshold=threshold)
        boot, ci = bootstrap_metrics(
            y_temp,
            temporal_prob,
            threshold=threshold,
            n_bootstrap=config.n_bootstrap,
            seed=stable_seed(task_seed, model_key + "_bootstrap"),
        )

        models[model_key] = {
            "label": model_label,
            "model": model,
            "params": params,
            "development_oof_prob": np.asarray(oof, dtype=float),
            "temporal_prob": np.asarray(temporal_prob, dtype=float),
            "threshold_info": threshold_info,
            "point": point,
            "bootstrap": boot,
            "ci": ci,
            "tuning_history": tuning,
        }

        joblib.dump(model, paths["models"] / f"{spec.key}_{model_key}.joblib")
        save_dataframe(
            boot,
            paths["bootstrap"] / f"{spec.key}_{model_key}_bootstrap.csv",
        )
        save_dataframe(
            tuning,
            paths["diagnostics"] / f"{spec.key}_{model_key}_tuning.csv",
        )

    dev_predictions = pd.DataFrame(
        {
            config.id_col: dev[config.id_col].to_numpy(),
            config.admission_col: dev[config.admission_col].to_numpy(),
            "period": "development_oof",
            "task": spec.key,
            "y_true": y_dev.to_numpy(),
            "lr_probability": models["lr"]["development_oof_prob"],
            "lgbm_probability": models["lgbm"]["development_oof_prob"],
        }
    )
    temp_predictions = pd.DataFrame(
        {
            config.id_col: temp[config.id_col].to_numpy(),
            config.admission_col: temp[config.admission_col].to_numpy(),
            "period": "temporal_evaluation",
            "task": spec.key,
            "y_true": y_temp.to_numpy(),
            "lr_probability": models["lr"]["temporal_prob"],
            "lgbm_probability": models["lgbm"]["temporal_prob"],
        }
    )
    for extra in [
        config.los_col,
        "Outcome_nonhome",
        "Outcome_transfer_no_death",
        "LOS_21",
        "LOS_28",
        "is_inhospital_death",
        "flag_day7_landmark",
    ]:
        if extra in temp.columns and extra not in temp_predictions.columns:
            temp_predictions[extra] = temp[extra].to_numpy()
        if extra in dev.columns and extra not in dev_predictions.columns:
            dev_predictions[extra] = dev[extra].to_numpy()

    save_dataframe(
        dev_predictions,
        paths["predictions"] / f"{spec.key}_development_oof_predictions.csv",
    )
    save_dataframe(
        temp_predictions,
        paths["predictions"] / f"{spec.key}_temporal_predictions.csv",
    )

    return {
        "spec": spec,
        "task_data": task_data,
        "development": dev,
        "temporal": temp,
        "y_dev": y_dev,
        "y_temp": y_temp,
        "x_dev": x_dev,
        "x_temp": x_temp,
        "development_predictions": dev_predictions,
        "temporal_predictions": temp_predictions,
        "models": models,
    }


def performance_row(
    task_result: dict[str, Any],
    model_key: str,
) -> dict[str, Any]:
    spec: TaskSpec = task_result["spec"]
    block = task_result["models"][model_key]
    point = block["point"]
    ci = block["ci"]
    row: dict[str, Any] = {
        "Analysis": spec.analysis_type,
        "Task": spec.key,
        "Landmark": f"Day {spec.landmark_day}",
        "Outcome": spec.outcome_label,
        "Model": block["label"],
        "N_development": len(task_result["development"]),
        "Events_development": int(task_result["y_dev"].sum()),
        "Event_rate_development": float(task_result["y_dev"].mean()),
        "N_temporal": len(task_result["temporal"]),
        "Events_temporal": int(task_result["y_temp"].sum()),
        "Event_rate_temporal": float(task_result["y_temp"].mean()),
        "Screening_threshold": block["threshold_info"]["threshold"],
        "OOF_sensitivity_at_threshold": block["threshold_info"][
            "development_oof_sensitivity"
        ],
        "OOF_specificity_at_threshold": block["threshold_info"][
            "development_oof_specificity"
        ],
    }
    for metric, value in point.items():
        row[metric] = value
        low, high = ci.get(metric, (np.nan, np.nan))
        row[f"{metric}_95CI_low"] = low
        row[f"{metric}_95CI_high"] = high
    return row

# %% 
# =============================================================================
# 7. Paired landmark comparison and Day 7 operational matrix
# =============================================================================


def paired_bootstrap_delta(
    y_true: np.ndarray,
    p_day3: np.ndarray,
    p_day7: np.ndarray,
    n_bootstrap: int,
    seed: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    y = np.asarray(y_true, dtype=int)
    p3 = np.asarray(p_day3, dtype=float)
    p7 = np.asarray(p_day7, dtype=float)
    point3 = calculate_metrics(y, p3)
    point7 = calculate_metrics(y, p7)
    metric_names = [
        "AUROC",
        "PR_AUC",
        "Brier",
        "Calibration_intercept",
        "Calibration_slope",
    ]
    point_delta = {f"Delta_{m}": point7[m] - point3[m] for m in metric_names}

    rng = np.random.default_rng(seed)
    n = len(y)
    rows: list[dict[str, float]] = []
    attempts = 0
    while len(rows) < n_bootstrap and attempts < n_bootstrap * 10:
        attempts += 1
        idx = rng.integers(0, n, size=n)
        if np.unique(y[idx]).size < 2:
            continue
        try:
            m3 = calculate_metrics(y[idx], p3[idx])
            m7 = calculate_metrics(y[idx], p7[idx])
        except Exception:
            continue
        rows.append({f"Delta_{m}": m7[m] - m3[m] for m in metric_names})
    boot = pd.DataFrame(rows)

    summary_rows: list[dict[str, Any]] = []
    for metric in metric_names:
        col = f"Delta_{metric}"
        values = boot[col].dropna().to_numpy()
        summary_rows.append(
            {
                "Metric": metric,
                "Day3": point3[metric],
                "Day7": point7[metric],
                "Delta_Day7_minus_Day3": point_delta[col],
                "Delta_95CI_low": float(np.percentile(values, 2.5)) if len(values) else np.nan,
                "Delta_95CI_high": float(np.percentile(values, 97.5)) if len(values) else np.nan,
                "Valid_bootstrap_resamples": len(values),
            }
        )
    return pd.DataFrame(summary_rows), boot


def paired_landmark_analysis(
    primary_results: dict[str, dict[str, Any]],
    config: AnalysisConfig,
    paths: dict[str, Path],
) -> tuple[pd.DataFrame, pd.DataFrame]:
    summaries: list[pd.DataFrame] = []
    all_boot: list[pd.DataFrame] = []
    comparisons = [
        ("nonhome", "day3_nonhome", "day7_nonhome"),
        ("los21", "day3_los21", "day7_los21"),
    ]
    for outcome_key, day3_key, day7_key in comparisons:
        r3 = primary_results[day3_key]
        r7 = primary_results[day7_key]
        for model_key in ["lr", "lgbm"]:
            d3 = r3["temporal_predictions"].copy()
            d7 = r7["temporal_predictions"][[config.id_col, "y_true", f"{model_key}_probability"]].copy()
            d3 = d3.loc[d3["flag_day7_landmark"].eq(1)].copy()
            d3 = d3[[config.id_col, "y_true", f"{model_key}_probability"]].rename(
                columns={
                    "y_true": "y_day3",
                    f"{model_key}_probability": "p_day3",
                }
            )
            d7 = d7.rename(
                columns={
                    "y_true": "y_day7",
                    f"{model_key}_probability": "p_day7",
                }
            )
            paired = d3.merge(d7, on=config.id_col, how="inner", validate="one_to_one")
            if not np.array_equal(paired["y_day3"].to_numpy(), paired["y_day7"].to_numpy()):
                raise AssertionError(f"Paired outcomes differ for {outcome_key}, {model_key}")
            summary, boot = paired_bootstrap_delta(
                paired["y_day3"].to_numpy(),
                paired["p_day3"].to_numpy(),
                paired["p_day7"].to_numpy(),
                n_bootstrap=config.n_bootstrap,
                seed=stable_seed(config.seed, f"paired_{outcome_key}_{model_key}"),
            )
            summary.insert(0, "Outcome", outcome_key)
            summary.insert(1, "Model", r3["models"][model_key]["label"])
            summary.insert(2, "N_paired", len(paired))
            boot.insert(0, "Outcome", outcome_key)
            boot.insert(1, "Model", r3["models"][model_key]["label"])
            summaries.append(summary)
            all_boot.append(boot)

    summary_table = pd.concat(summaries, ignore_index=True)
    bootstrap_table = pd.concat(all_boot, ignore_index=True)
    save_dataframe(summary_table, paths["tables"] / "paired_day3_day7_summary.csv")
    save_dataframe(bootstrap_table, paths["bootstrap"] / "paired_day3_day7_bootstrap.csv")
    return summary_table, bootstrap_table



def build_landmark_matrix(
    primary_results: dict[str, dict[str, Any]],
    model_key: str,
    config: AnalysisConfig,
    landmark_day: int,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build the two-dimensional operational risk matrix at Day 3 or Day 7.

    The same development-derived, model-specific screening thresholds used
    in the primary analyses are applied at each landmark.
    """
    if landmark_day not in {3, 7}:
        raise ValueError(
            f"landmark_day must be 3 or 7; got {landmark_day}"
        )

    nonhome_key = f"day{landmark_day}_nonhome"
    los21_key = f"day{landmark_day}_los21"

    if nonhome_key not in primary_results or los21_key not in primary_results:
        raise KeyError(
            f"Missing landmark tasks: {nonhome_key}, {los21_key}"
        )

    nonhome = primary_results[nonhome_key]
    los21 = primary_results[los21_key]

    a = nonhome["temporal_predictions"][
        [
            config.id_col,
            config.los_col,
            "Outcome_nonhome",
            "LOS_21",
            f"{model_key}_probability",
        ]
    ].copy()

    a = a.rename(
        columns={
            f"{model_key}_probability": "p_nonhome"
        }
    )

    b = los21["temporal_predictions"][
        [
            config.id_col,
            f"{model_key}_probability",
        ]
    ].copy()

    b = b.rename(
        columns={
            f"{model_key}_probability": "p_los21"
        }
    )

    matrix = a.merge(
        b,
        on=config.id_col,
        how="inner",
        validate="one_to_one",
    )

    th_nonhome = float(
        nonhome["models"][model_key]["threshold_info"]["threshold"]
    )
    th_los = float(
        los21["models"][model_key]["threshold_info"]["threshold"]
    )

    matrix["high_nonhome"] = (
        matrix["p_nonhome"]
        .ge(th_nonhome)
        .astype("int8")
    )
    matrix["high_los21"] = (
        matrix["p_los21"]
        .ge(th_los)
        .astype("int8")
    )

    matrix["matrix_group"] = np.select(
        [
            matrix["high_nonhome"].eq(0)
            & matrix["high_los21"].eq(0),

            matrix["high_nonhome"].eq(1)
            & matrix["high_los21"].eq(0),

            matrix["high_nonhome"].eq(0)
            & matrix["high_los21"].eq(1),

            matrix["high_nonhome"].eq(1)
            & matrix["high_los21"].eq(1),
        ],
        [
            "Lower risk for both outcomes",
            "Higher non-home risk only",
            "Higher prolonged-hospitalization risk only",
            "Higher risk for both outcomes",
        ],
        default="Unclassified",
    )

    matrix["observed_pattern"] = np.select(
        [
            matrix["Outcome_nonhome"].eq(0)
            & matrix["LOS_21"].eq(0),

            matrix["Outcome_nonhome"].eq(1)
            & matrix["LOS_21"].eq(0),

            matrix["Outcome_nonhome"].eq(0)
            & matrix["LOS_21"].eq(1),

            matrix["Outcome_nonhome"].eq(1)
            & matrix["LOS_21"].eq(1),
        ],
        [
            "Neither outcome",
            "Non-home discharge only",
            "Prolonged hospitalization only",
            "Both outcomes",
        ],
        default="Unclassified",
    )

    matrix["Landmark"] = f"Day {landmark_day}"
    matrix["Model"] = nonhome["models"][model_key]["label"]

    order = [
        "Lower risk for both outcomes",
        "Higher non-home risk only",
        "Higher prolonged-hospitalization risk only",
        "Higher risk for both outcomes",
    ]

    summary = (
        matrix
        .groupby(
            "matrix_group",
            observed=True,
        )
        .agg(
            n=(config.id_col, "size"),
            observed_nonhome_n=("Outcome_nonhome", "sum"),
            observed_nonhome_rate=("Outcome_nonhome", "mean"),
            observed_los21_n=("LOS_21", "sum"),
            observed_los21_rate=("LOS_21", "mean"),
            median_LOS=(config.los_col, "median"),
            q1_LOS=(
                config.los_col,
                lambda x: x.quantile(0.25),
            ),
            q3_LOS=(
                config.los_col,
                lambda x: x.quantile(0.75),
            ),
            mean_predicted_nonhome=("p_nonhome", "mean"),
            mean_predicted_los21=("p_los21", "mean"),
        )
        .reindex(order)
        .reset_index()
    )

    summary["cohort_percent"] = (
        summary["n"]
        / len(matrix)
    )
    summary["Landmark"] = f"Day {landmark_day}"
    summary["Model"] = nonhome["models"][model_key]["label"]
    summary["threshold_nonhome"] = th_nonhome
    summary["threshold_los21"] = th_los

    return matrix, summary


def build_day7_matrix(
    primary_results: dict[str, dict[str, Any]],
    model_key: str,
    config: AnalysisConfig,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Backward-compatible wrapper retained for older code paths.
    """
    return build_landmark_matrix(
        primary_results=primary_results,
        model_key=model_key,
        config=config,
        landmark_day=7,
    )


# %% 
# =============================================================================
# 8. Characteristics, missingness, coefficients, and SHAP summaries
# =============================================================================


def smd_continuous(a: pd.Series, b: pd.Series) -> float:
    x = pd.to_numeric(a, errors="coerce").dropna().to_numpy(dtype=float)
    y = pd.to_numeric(b, errors="coerce").dropna().to_numpy(dtype=float)
    if len(x) < 2 or len(y) < 2:
        return np.nan
    pooled = math.sqrt((np.var(x, ddof=1) + np.var(y, ddof=1)) / 2)
    return float((np.mean(y) - np.mean(x)) / pooled) if pooled > 0 else 0.0


def smd_binary(a: pd.Series, b: pd.Series) -> float:
    x = pd.to_numeric(a, errors="coerce").dropna().to_numpy(dtype=float)
    y = pd.to_numeric(b, errors="coerce").dropna().to_numpy(dtype=float)
    if len(x) == 0 or len(y) == 0:
        return np.nan
    p1, p2 = float(np.mean(x)), float(np.mean(y))
    denom = math.sqrt((p1 * (1 - p1) + p2 * (1 - p2)) / 2)
    return float((p2 - p1) / denom) if denom > 0 else 0.0


def fmt_continuous(series: pd.Series) -> str:
    x = pd.to_numeric(series, errors="coerce").dropna()
    if x.empty:
        return "NA"
    return f"{x.median():.2f} [{x.quantile(0.25):.2f}, {x.quantile(0.75):.2f}]"


def fmt_binary(series: pd.Series) -> str:
    x = pd.to_numeric(series, errors="coerce").dropna()
    if x.empty:
        return "NA"
    n = int(x.eq(1).sum())
    return f"{n} ({100*n/len(x):.1f}%)"


def make_characteristics_table(
    data: pd.DataFrame,
    landmark_day: int,
    continuous_features: tuple[str, ...],
    binary_features: tuple[str, ...],
    config: AnalysisConfig,
) -> pd.DataFrame:
    flag = f"flag_day{landmark_day}_landmark"
    cohort = data.loc[data[flag].eq(1)].copy()
    dev = cohort.loc[cohort["period"].eq("development")]
    temp = cohort.loc[cohort["period"].eq("temporal_evaluation")]
    rows: list[dict[str, Any]] = []

    # Outcomes and LOS first.
    rows.extend(
        [
            {
                "Domain": "Outcomes",
                "Variable": "No direct home discharge",
                f"Development (n={len(dev)})": fmt_binary(dev["Outcome_nonhome"]),
                f"Temporal evaluation (n={len(temp)})": fmt_binary(temp["Outcome_nonhome"]),
                "SMD": smd_binary(dev["Outcome_nonhome"], temp["Outcome_nonhome"]),
                "Missing_development_n": int(dev["Outcome_nonhome"].isna().sum()),
                "Missing_temporal_n": int(temp["Outcome_nonhome"].isna().sum()),
            },
            {
                "Domain": "Outcomes",
                "Variable": "Hospital length of stay >21 days",
                f"Development (n={len(dev)})": fmt_binary(dev["LOS_21"]),
                f"Temporal evaluation (n={len(temp)})": fmt_binary(temp["LOS_21"]),
                "SMD": smd_binary(dev["LOS_21"], temp["LOS_21"]),
                "Missing_development_n": 0,
                "Missing_temporal_n": 0,
            },
            {
                "Domain": "Outcomes",
                "Variable": "Hospital length of stay, days",
                f"Development (n={len(dev)})": fmt_continuous(dev[config.los_col]),
                f"Temporal evaluation (n={len(temp)})": fmt_continuous(temp[config.los_col]),
                "SMD": smd_continuous(dev[config.los_col], temp[config.los_col]),
                "Missing_development_n": int(dev[config.los_col].isna().sum()),
                "Missing_temporal_n": int(temp[config.los_col].isna().sum()),
            },
        ]
    )

    for col in continuous_features:
        rows.append(
            {
                "Domain": "Predictor",
                "Variable": PUBLICATION_LABELS.get(col, col),
                f"Development (n={len(dev)})": fmt_continuous(dev[col]),
                f"Temporal evaluation (n={len(temp)})": fmt_continuous(temp[col]),
                "SMD": smd_continuous(dev[col], temp[col]),
                "Missing_development_n": int(dev[col].isna().sum()),
                "Missing_temporal_n": int(temp[col].isna().sum()),
            }
        )
    for col in binary_features:
        rows.append(
            {
                "Domain": "Predictor",
                "Variable": PUBLICATION_LABELS.get(col, col),
                f"Development (n={len(dev)})": fmt_binary(dev[col]),
                f"Temporal evaluation (n={len(temp)})": fmt_binary(temp[col]),
                "SMD": smd_binary(dev[col], temp[col]),
                "Missing_development_n": int(dev[col].isna().sum()),
                "Missing_temporal_n": int(temp[col].isna().sum()),
            }
        )
    return pd.DataFrame(rows)


def make_missingness_table(
    data: pd.DataFrame,
    landmark_day: int,
    features: list[str],
) -> pd.DataFrame:
    flag = f"flag_day{landmark_day}_landmark"
    rows: list[dict[str, Any]] = []
    for period in ["development", "temporal_evaluation"]:
        sub = data.loc[data[flag].eq(1) & data["period"].eq(period)]
        for col in features:
            n_missing = int(sub[col].isna().sum())
            rows.append(
                {
                    "Landmark": f"Day {landmark_day}",
                    "Period": period,
                    "Raw_variable": col,
                    "Variable": PUBLICATION_LABELS.get(col, col),
                    "N": len(sub),
                    "Missing_n": n_missing,
                    "Missing_percent": 100 * n_missing / len(sub) if len(sub) else np.nan,
                }
            )
    return pd.DataFrame(rows)


def clean_feature_name(name: str) -> str:
    return name.split("__", 1)[1] if "__" in name else name


def extract_lr_coefficients(
    task_result: dict[str, Any],
) -> pd.DataFrame:
    spec: TaskSpec = task_result["spec"]
    pipeline: Pipeline = task_result["models"]["lr"]["model"]
    preprocess: ColumnTransformer = pipeline.named_steps["preprocess"]
    model: LogisticRegression = pipeline.named_steps["model"]
    names = [clean_feature_name(x) for x in preprocess.get_feature_names_out()]
    coefs = model.coef_.ravel()
    if len(names) != len(coefs):
        raise AssertionError("Coefficient and feature-name lengths differ")
    feature_type = [
        "continuous_standardized" if n in set(spec.continuous_features) else "binary"
        for n in names
    ]
    out = pd.DataFrame(
        {
            "Task": spec.key,
            "Landmark": f"Day {spec.landmark_day}",
            "Outcome": spec.outcome_label,
            "Raw_variable": names,
            "Variable": [PUBLICATION_LABELS.get(n, n) for n in names],
            "Feature_type": feature_type,
            "Coefficient": coefs,
            "Odds_ratio_per_model_unit": np.exp(coefs),
            "Absolute_coefficient": np.abs(coefs),
        }
    ).sort_values("Absolute_coefficient", ascending=False)
    intercept = pd.DataFrame(
        {
            "Task": [spec.key],
            "Landmark": [f"Day {spec.landmark_day}"],
            "Outcome": [spec.outcome_label],
            "Raw_variable": ["(Intercept)"],
            "Variable": ["Intercept"],
            "Feature_type": ["intercept"],
            "Coefficient": [float(model.intercept_[0])],
            "Odds_ratio_per_model_unit": [float(np.exp(model.intercept_[0]))],
            "Absolute_coefficient": [abs(float(model.intercept_[0]))],
        }
    )
    return pd.concat([intercept, out], ignore_index=True)


def extract_lgbm_importance(task_result: dict[str, Any]) -> pd.DataFrame:
    spec: TaskSpec = task_result["spec"]
    pipeline: Pipeline = task_result["models"]["lgbm"]["model"]
    preprocess: ColumnTransformer = pipeline.named_steps["preprocess"]
    model: LGBMClassifier = pipeline.named_steps["model"]
    names = [clean_feature_name(x) for x in preprocess.get_feature_names_out()]
    gain = model.booster_.feature_importance(importance_type="gain")
    split = model.booster_.feature_importance(importance_type="split")
    out = pd.DataFrame(
        {
            "Task": spec.key,
            "Landmark": f"Day {spec.landmark_day}",
            "Outcome": spec.outcome_label,
            "Raw_variable": names,
            "Variable": [PUBLICATION_LABELS.get(n, n) for n in names],
            "Gain_importance": gain,
            "Split_importance": split,
        }
    )
    gain_sum = out["Gain_importance"].sum()
    out["Gain_fraction"] = out["Gain_importance"] / gain_sum if gain_sum > 0 else np.nan
    return out.sort_values("Gain_importance", ascending=False)


def calculate_shap_summary(
    task_result: dict[str, Any],
    config: AnalysisConfig,
    paths: dict[str, Path],
) -> pd.DataFrame:
    # Import SHAP only when explicitly requested. Eager SHAP/Numba loading can
    # hard-crash some Windows Jupyter kernels before Python can raise an exception.
    import shap

    spec: TaskSpec = task_result["spec"]
    pipeline: Pipeline = task_result["models"]["lgbm"]["model"]
    preprocess: ColumnTransformer = pipeline.named_steps["preprocess"]
    model: LGBMClassifier = pipeline.named_steps["model"]
    x = task_result["x_temp"].copy()
    if len(x) > config.shap_max_patients:
        x = x.sample(
            n=config.shap_max_patients,
            random_state=stable_seed(config.seed, spec.key + "_shap_sample"),
        )
    x_transformed = preprocess.transform(x)
    if hasattr(x_transformed, "toarray"):
        x_transformed = x_transformed.toarray()
    names = [clean_feature_name(v) for v in preprocess.get_feature_names_out()]

    explainer = shap.TreeExplainer(model)
    explanation = explainer(x_transformed)
    values = np.asarray(explanation.values)
    if values.ndim == 3:
        values = values[:, :, -1]
    if values.shape[1] != len(names):
        raise AssertionError("SHAP feature dimension does not match feature names")

    summary = pd.DataFrame(
        {
            "Task": spec.key,
            "Landmark": f"Day {spec.landmark_day}",
            "Outcome": spec.outcome_label,
            "Raw_variable": names,
            "Variable": [PUBLICATION_LABELS.get(n, n) for n in names],
            "Mean_absolute_SHAP": np.abs(values).mean(axis=0),
            "Mean_SHAP": values.mean(axis=0),
        }
    ).sort_values("Mean_absolute_SHAP", ascending=False)

    # SHAP plot uses transformed values and publication labels.
    display_names = [PUBLICATION_LABELS.get(n, n) for n in names]
    plt.figure(figsize=(8.0, 6.5))
    shap.summary_plot(
        values,
        x_transformed,
        feature_names=display_names,
        max_display=15,
        show=False,
        plot_size=None,
    )
    fig = plt.gcf()
    fig.tight_layout()
    save_figure(
        fig,
        paths["figures"] / f"Figure_S_SHAP_{spec.key}",
        config.figure_dpi,
    )
    return summary

# %% 
# =============================================================================
# 9. Publication figures
# =============================================================================


def configure_matplotlib(config: AnalysisConfig) -> None:
    matplotlib.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": [config.figure_font, "DejaVu Serif"],
            "font.size": 10,
            "axes.titlesize": 11,
            "axes.labelsize": 10,
            "xtick.labelsize": 9,
            "ytick.labelsize": 9,
            "legend.fontsize": 9,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "axes.spines.top": False,
            "axes.spines.right": False,
        }
    )


def panel_specs() -> list[tuple[str, str]]:
    return [
        ("day3_nonhome", "A  Day 3: no direct home discharge"),
        ("day3_los21", "B  Day 3: hospital stay >21 days"),
        ("day7_nonhome", "C  Day 7: no direct home discharge"),
        ("day7_los21", "D  Day 7: hospital stay >21 days"),
    ]


def plot_roc_panels(primary_results: dict[str, dict[str, Any]]) -> plt.Figure:
    fig, axes = plt.subplots(2, 2, figsize=(10.0, 8.0))
    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            p = block["temporal_prob"]
            fpr, tpr, _ = roc_curve(y, p)
            auc_value = block["point"]["AUROC"]
            low, high = block["ci"]["AUROC"]
            ax.plot(fpr, tpr, label=f"{block['label']} {auc_value:.3f} ({low:.3f}–{high:.3f})")
        ax.plot([0, 1], [0, 1], linestyle="--", linewidth=0.9)
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlabel("1 − specificity")
        ax.set_ylabel("Sensitivity")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.legend(frameon=False, loc="lower right")
    fig.tight_layout()
    return fig


def plot_pr_panels(primary_results: dict[str, dict[str, Any]]) -> plt.Figure:
    fig, axes = plt.subplots(2, 2, figsize=(10.0, 8.0))
    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        prevalence = float(y.mean())
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            precision, recall, _ = precision_recall_curve(y, block["temporal_prob"])
            value = block["point"]["PR_AUC"]
            low, high = block["ci"]["PR_AUC"]
            ax.plot(recall, precision, label=f"{block['label']} {value:.3f} ({low:.3f}–{high:.3f})")
        ax.axhline(prevalence, linestyle="--", linewidth=0.9, label=f"Event rate {prevalence:.3f}")
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlabel("Recall")
        ax.set_ylabel("Precision")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.legend(frameon=False, loc="best")
    fig.tight_layout()
    return fig


def plot_calibration_panels(primary_results: dict[str, dict[str, Any]]) -> plt.Figure:
    fig, axes = plt.subplots(2, 2, figsize=(10.0, 8.0))
    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            cal = calibration_curve_quantile(y, block["temporal_prob"], n_bins=10)
            intercept = block["point"]["Calibration_intercept"]
            slope = block["point"]["Calibration_slope"]
            ax.plot(
                cal["mean_predicted"],
                cal["observed"],
                marker="o",
                label=f"{block['label']} (int {intercept:.2f}, slope {slope:.2f})",
            )
        ax.plot([0, 1], [0, 1], linestyle="--", linewidth=0.9)
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlabel("Mean predicted probability")
        ax.set_ylabel("Observed proportion")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.legend(frameon=False, loc="best")
    fig.tight_layout()
    return fig


def plot_dca_panels(
    primary_results: dict[str, dict[str, Any]],
    config: AnalysisConfig,
) -> plt.Figure:
    thresholds = np.arange(
        config.dca_threshold_min,
        config.dca_threshold_max + config.dca_threshold_step / 2,
        config.dca_threshold_step,
    )
    fig, axes = plt.subplots(2, 2, figsize=(10.0, 8.0))
    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        references_plotted = False
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            dca = decision_curve(y, block["temporal_prob"], thresholds)
            ax.plot(dca["threshold"], dca["model"], label=block["label"])
            if not references_plotted:
                ax.plot(dca["threshold"], dca["treat_all"], linestyle="--", label="Treat all")
                ax.plot(dca["threshold"], dca["treat_none"], linestyle=":", label="Treat none")
                references_plotted = True
        ax.axhline(0, linewidth=0.7)
        ax.set_title(title, loc="left", fontweight="bold")
        ax.set_xlabel("Threshold probability")
        ax.set_ylabel("Net benefit")
        ax.set_xlim(config.dca_threshold_min, config.dca_threshold_max)
        ax.legend(frameon=False, loc="best")
    fig.tight_layout()
    return fig


def plot_day7_matrix(
    matrix: pd.DataFrame,
    summary: pd.DataFrame,
    model_label: str,
) -> plt.Figure:
    fig, ax = plt.subplots(figsize=(8.0, 7.0))
    patterns = [
        ("Neither outcome", "o"),
        ("Non-home discharge only", "^"),
        ("Prolonged hospitalization only", "s"),
        ("Both outcomes", "D"),
    ]
    for pattern, marker in patterns:
        sub = matrix.loc[matrix["observed_pattern"].eq(pattern)]
        ax.scatter(
            sub["p_nonhome"],
            sub["p_los21"],
            marker=marker,
            alpha=0.55,
            s=25,
            label=pattern,
        )
    th_x = float(summary["threshold_nonhome"].iloc[0])
    th_y = float(summary["threshold_los21"].iloc[0])
    ax.axvline(th_x, linestyle="--", linewidth=1.0)
    ax.axhline(th_y, linestyle="--", linewidth=1.0)
    ax.set_xlabel("Predicted probability of no direct home discharge")
    ax.set_ylabel("Predicted probability of hospital stay >21 days")
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)

    annotations = {
        "Lower risk for both outcomes": (th_x / 2, th_y / 2),
        "Higher non-home risk only": ((th_x + 1) / 2, th_y / 2),
        "Higher prolonged-hospitalization risk only": (th_x / 2, (th_y + 1) / 2),
        "Higher risk for both outcomes": ((th_x + 1) / 2, (th_y + 1) / 2),
    }
    for _, row in summary.iterrows():
        x, y = annotations[row["matrix_group"]]
        ax.text(
            x,
            y,
            f"n={int(row['n'])}\nNon-home {100*row['observed_nonhome_rate']:.1f}%\n"
            f">21 d {100*row['observed_los21_rate']:.1f}%",
            ha="center",
            va="center",
            bbox={"facecolor": "white", "edgecolor": "black", "alpha": 0.85, "pad": 4},
        )
    ax.text(0.01, 1.02, model_label, transform=ax.transAxes, fontweight="bold", ha="left")
    ax.legend(frameon=False, loc="upper left", bbox_to_anchor=(1.02, 1.0))
    fig.tight_layout()
    return fig


def plot_lr_coefficients(
    coefficients: pd.DataFrame,
    top_n: int = 12,
) -> plt.Figure:
    tasks = [key for key, _ in panel_specs()]
    titles = dict(panel_specs())
    fig, axes = plt.subplots(2, 2, figsize=(11.0, 9.0))
    for ax, task in zip(axes.ravel(), tasks):
        sub = coefficients.loc[
            coefficients["Task"].eq(task) & ~coefficients["Raw_variable"].eq("(Intercept)")
        ].nlargest(top_n, "Absolute_coefficient")
        sub = sub.sort_values("Coefficient")
        ax.barh(sub["Variable"], sub["Coefficient"])
        ax.axvline(0, linewidth=0.8)
        ax.set_title(titles[task], loc="left", fontweight="bold")
        ax.set_xlabel("Penalized logistic regression coefficient")
    fig.tight_layout()
    return fig

# %% 

# =============================================================================
# 9B. FINAL BMC MIDM publication figure functions
# =============================================================================

def mm_to_in(mm: float) -> float:
    return mm / 25.4


def configure_matplotlib(config: AnalysisConfig) -> None:
    from matplotlib import font_manager

    available = {font.name for font in font_manager.fontManager.ttflist}
    preferred = config.figure_font if config.figure_font in available else "DejaVu Serif"

    matplotlib.rcParams.update(
        {
            "font.family": preferred,
            "font.size": 8.0,
            "axes.titlesize": 8.6,
            "axes.labelsize": 8.0,
            "xtick.labelsize": 7.2,
            "ytick.labelsize": 7.2,
            "legend.fontsize": 6.9,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.linewidth": 0.7,
            "lines.linewidth": 1.25,
            "savefig.transparent": False,
        }
    )


# Fixed, color-blind–friendly model palette.
_MODEL_STYLE = {
    "lr": {
        "color": "#0072B2",
        "linestyle": "-",
        "marker": "o",
        "short": "Penalized logistic regression",
    },
    "lgbm": {
        "color": "#D55E00",
        "linestyle": "--",
        "marker": "s",
        "short": "LightGBM",
    },
}

_REFERENCE_STYLE = {
    "ideal": {"color": "#8C8C8C", "linestyle": ":"},
    "treat_all": {"color": "#595959", "linestyle": "-."},
    "treat_none": {"color": "#A6A6A6", "linestyle": ":"},
}

_MATRIX_COLORS = {
    "Lower risk for both outcomes": "#E2F0D9",
    "Higher non-home risk only": "#D9EAF7",
    "Higher prolonged-hospitalization risk only": "#FFF2CC",
    "Higher risk for both outcomes": "#F4CCCC",
}


def panel_specs() -> list[tuple[str, str]]:
    return [
        ("day3_nonhome", "(A) Day 3: no direct home discharge"),
        ("day3_los21", "(B) Day 3: hospital stay >21 days"),
        ("day7_nonhome", "(C) Day 7: no direct home discharge"),
        ("day7_los21", "(D) Day 7: hospital stay >21 days"),
    ]


def _format_probability_axes(ax: plt.Axes) -> None:
    ticks = np.linspace(0, 1, 6)
    ax.set_xticks(ticks)
    ax.set_yticks(ticks)
    ax.tick_params(direction="out", length=2.5, width=0.6)


def _add_metric_box(
    ax: plt.Axes,
    lines: list[str],
    *,
    loc: str = "lower right",
    fontsize: float = 6.6,
) -> None:
    positions = {
        "lower right": (0.98, 0.03, "right", "bottom"),
        "lower left": (0.02, 0.03, "left", "bottom"),
        "upper right": (0.98, 0.97, "right", "top"),
        "upper left": (0.02, 0.97, "left", "top"),
    }
    x, y, ha, va = positions[loc]
    ax.text(
        x,
        y,
        "\n".join(lines),
        transform=ax.transAxes,
        ha=ha,
        va=va,
        fontsize=fontsize,
        bbox={
            "facecolor": "white",
            "edgecolor": "#D0D0D0",
            "alpha": 0.90,
            "pad": 2.0,
        },
        zorder=5,
    )


def _common_model_legend(fig: plt.Figure, *, include_references: bool = False) -> None:
    from matplotlib.lines import Line2D

    handles = [
        Line2D(
            [0], [0],
            color=_MODEL_STYLE["lr"]["color"],
            linestyle=_MODEL_STYLE["lr"]["linestyle"],
            linewidth=1.5,
            label=_MODEL_STYLE["lr"]["short"],
        ),
        Line2D(
            [0], [0],
            color=_MODEL_STYLE["lgbm"]["color"],
            linestyle=_MODEL_STYLE["lgbm"]["linestyle"],
            linewidth=1.5,
            label=_MODEL_STYLE["lgbm"]["short"],
        ),
    ]
    if include_references:
        handles.extend(
            [
                Line2D(
                    [0], [0],
                    color=_REFERENCE_STYLE["treat_all"]["color"],
                    linestyle=_REFERENCE_STYLE["treat_all"]["linestyle"],
                    linewidth=1.2,
                    label="Treat all",
                ),
                Line2D(
                    [0], [0],
                    color=_REFERENCE_STYLE["treat_none"]["color"],
                    linestyle=_REFERENCE_STYLE["treat_none"]["linestyle"],
                    linewidth=1.2,
                    label="Treat none",
                ),
            ]
        )

    fig.legend(
        handles=handles,
        loc="lower center",
        bbox_to_anchor=(0.5, 0.005),
        ncol=len(handles),
        frameon=False,
        handlelength=2.6,
        columnspacing=1.5,
    )


def plot_roc_panels(primary_results: dict[str, dict[str, Any]]) -> plt.Figure:
    fig, axes = plt.subplots(
        2, 2,
        figsize=(mm_to_in(170), mm_to_in(160)),
    )

    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()

        metric_lines = []
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            p = block["temporal_prob"]
            fpr, tpr, _ = roc_curve(y, p)
            value = block["point"]["AUROC"]
            low, high = block["ci"]["AUROC"]
            st = _MODEL_STYLE[model_key]

            ax.plot(
                fpr,
                tpr,
                color=st["color"],
                linestyle=st["linestyle"],
                linewidth=1.35,
            )
            metric_lines.append(
                f"{'LR' if model_key == 'lr' else 'LightGBM'}: "
                f"{value:.3f} ({low:.3f}–{high:.3f})"
            )

        ax.plot(
            [0, 1], [0, 1],
            color=_REFERENCE_STYLE["ideal"]["color"],
            linestyle=_REFERENCE_STYLE["ideal"]["linestyle"],
            linewidth=0.9,
        )
        ax.set_title(title, loc="left", fontweight="bold", pad=4)
        ax.set_xlabel("1 − specificity")
        ax.set_ylabel("Sensitivity")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        _format_probability_axes(ax)
        _add_metric_box(ax, metric_lines, loc="lower right", fontsize=6.5)

    _common_model_legend(fig)
    fig.subplots_adjust(
        left=0.10, right=0.99, top=0.98, bottom=0.13,
        wspace=0.28, hspace=0.34,
    )
    return fig


def plot_pr_panels(primary_results: dict[str, dict[str, Any]]) -> plt.Figure:
    fig, axes = plt.subplots(
        2, 2,
        figsize=(mm_to_in(170), mm_to_in(160)),
    )

    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        prevalence = float(y.mean())

        metric_lines = []
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            precision, recall, _ = precision_recall_curve(
                y,
                block["temporal_prob"],
            )
            value = block["point"]["PR_AUC"]
            low, high = block["ci"]["PR_AUC"]
            st = _MODEL_STYLE[model_key]

            ax.plot(
                recall,
                precision,
                color=st["color"],
                linestyle=st["linestyle"],
                linewidth=1.35,
            )
            metric_lines.append(
                f"{'LR' if model_key == 'lr' else 'LightGBM'}: "
                f"{value:.3f} ({low:.3f}–{high:.3f})"
            )

        ax.axhline(
            prevalence,
            color=_REFERENCE_STYLE["ideal"]["color"],
            linestyle=_REFERENCE_STYLE["ideal"]["linestyle"],
            linewidth=0.9,
        )
        metric_lines.append(f"Event rate: {prevalence:.3f}")

        ax.set_title(title, loc="left", fontweight="bold", pad=4)
        ax.set_xlabel("Recall")
        ax.set_ylabel("Precision")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        _format_probability_axes(ax)
        _add_metric_box(ax, metric_lines, loc="lower left", fontsize=6.4)

    _common_model_legend(fig)
    fig.subplots_adjust(
        left=0.10, right=0.99, top=0.98, bottom=0.13,
        wspace=0.28, hspace=0.34,
    )
    return fig


def plot_calibration_panels(
    primary_results: dict[str, dict[str, Any]]
) -> plt.Figure:
    """
    Calibration legend contains model names only.
    Calibration intercept and slope are shown in a compact panel annotation.
    """
    fig, axes = plt.subplots(
        2, 2,
        figsize=(mm_to_in(170), mm_to_in(160)),
    )

    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()

        metric_lines = []
        for model_key in ["lr", "lgbm"]:
            block = r["models"][model_key]
            cal = calibration_curve_quantile(
                y,
                block["temporal_prob"],
                n_bins=10,
            )
            intercept = block["point"]["Calibration_intercept"]
            slope = block["point"]["Calibration_slope"]
            st = _MODEL_STYLE[model_key]

            ax.plot(
                cal["mean_predicted"],
                cal["observed"],
                color=st["color"],
                linestyle=st["linestyle"],
                marker=st["marker"],
                markersize=3.0,
                linewidth=1.25,
                label=st["short"],
            )

            metric_lines.append(
                f"{'LR' if model_key == 'lr' else 'LightGBM'}: "
                f"intercept {intercept:.2f}, slope {slope:.2f}"
            )

        ax.plot(
            [0, 1], [0, 1],
            color=_REFERENCE_STYLE["ideal"]["color"],
            linestyle=_REFERENCE_STYLE["ideal"]["linestyle"],
            linewidth=0.9,
        )
        ax.set_title(title, loc="left", fontweight="bold", pad=4)
        ax.set_xlabel("Mean predicted probability")
        ax.set_ylabel("Observed proportion")
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        _format_probability_axes(ax)

        # Model names only in legend.
        ax.legend(
            frameon=False,
            loc="upper left",
            handlelength=2.4,
        )

        # Calibration metrics separated from legend.
        _add_metric_box(
            ax,
            metric_lines,
            loc="lower right",
            fontsize=6.25,
        )

    fig.subplots_adjust(
        left=0.10, right=0.99, top=0.98, bottom=0.09,
        wspace=0.28, hspace=0.34,
    )
    return fig


def plot_lr_coefficients(
    coefficients: pd.DataFrame,
    top_n: int = 10,
) -> plt.Figure:
    tasks = [key for key, _ in panel_specs()]
    titles = dict(panel_specs())

    fig, axes = plt.subplots(
        2, 2,
        figsize=(mm_to_in(170), mm_to_in(190)),
    )

    for ax, task in zip(axes.ravel(), tasks):
        sub = coefficients.loc[
            coefficients["Task"].eq(task)
            & ~coefficients["Raw_variable"].eq("(Intercept)")
        ].nlargest(top_n, "Absolute_coefficient")

        sub = sub.sort_values("Coefficient")

        ax.barh(
            sub["Variable"],
            sub["Coefficient"],
            color="#4C78A8",
            edgecolor="#2F4B5C",
            linewidth=0.35,
        )
        ax.axvline(0, color="#555555", linewidth=0.7)
        ax.set_title(
            titles[task],
            loc="left",
            fontweight="bold",
            pad=4,
        )
        ax.set_xlabel("Penalized logistic regression coefficient")
        ax.tick_params(axis="y", labelsize=6.2)

    # More left margin for long variable names.
    fig.subplots_adjust(
        left=0.30, right=0.99, top=0.98, bottom=0.08,
        wspace=0.72, hspace=0.38,
    )
    return fig

# %% 
# =============================================================================
# 10. Main analysis orchestration and export
# =============================================================================


def flatten_hyperparameters(
    results: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for task_key, result in results.items():
        spec: TaskSpec = result["spec"]
        for model_key, block in result["models"].items():
            base = {
                "Analysis": spec.analysis_type,
                "Task": task_key,
                "Model": block["label"],
            }
            for name, value in block["params"].items():
                rows.append({**base, "Parameter": name, "Value": value})
    return pd.DataFrame(rows)


def extract_full_model_parameters(
    results: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    """Export fitted estimator parameters, including fixed non-tuned settings."""
    rows: list[dict[str, Any]] = []
    for task_key, result in results.items():
        spec: TaskSpec = result["spec"]
        for model_key, block in result["models"].items():
            fitted_model = block["model"].named_steps["model"]
            parameters = fitted_model.get_params(deep=False)
            for name, value in sorted(parameters.items()):
                rows.append(
                    {
                        "Analysis": spec.analysis_type,
                        "Task": task_key,
                        "Model": block["label"],
                        "Parameter": name,
                        "Value": json.dumps(json_safe(value), ensure_ascii=False)
                        if isinstance(value, (dict, list, tuple))
                        else value,
                    }
                )
    return pd.DataFrame(rows)


def make_development_oof_performance(
    results: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for task_key, result in results.items():
        spec: TaskSpec = result["spec"]
        y = result["y_dev"].to_numpy(dtype=int)
        for model_key, block in result["models"].items():
            threshold = block["threshold_info"]["threshold"]
            metrics = calculate_metrics(
                y,
                block["development_oof_prob"],
                threshold=threshold,
            )
            rows.append(
                {
                    "Analysis": spec.analysis_type,
                    "Task": task_key,
                    "Landmark": f"Day {spec.landmark_day}",
                    "Outcome": spec.outcome_label,
                    "Model": block["label"],
                    "N": len(y),
                    "Events": int(y.sum()),
                    "Event_rate": float(y.mean()),
                    "Screening_threshold": threshold,
                    **metrics,
                }
            )
    return pd.DataFrame(rows)


def format_estimate_ci(
    row: pd.Series,
    metric: str,
    digits: int = 3,
) -> str:
    estimate = row.get(metric, np.nan)
    low = row.get(f"{metric}_95CI_low", np.nan)
    high = row.get(f"{metric}_95CI_high", np.nan)
    if not np.isfinite(estimate):
        return "NA"
    if not (np.isfinite(low) and np.isfinite(high)):
        return f"{estimate:.{digits}f}"
    return f"{estimate:.{digits}f} ({low:.{digits}f}–{high:.{digits}f})"


def make_publication_performance_table(
    performance: pd.DataFrame,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for _, row in performance.iterrows():
        rows.append(
            {
                "Landmark": row["Landmark"],
                "Outcome": row["Outcome"],
                "Model": row["Model"],
                "Temporal N": int(row["N_temporal"]),
                "Events": int(row["Events_temporal"]),
                "Event rate, %": round(100 * row["Event_rate_temporal"], 1),
                "AUROC (95% CI)": format_estimate_ci(row, "AUROC"),
                "PR-AUC (95% CI)": format_estimate_ci(row, "PR_AUC"),
                "Brier score (95% CI)": format_estimate_ci(row, "Brier"),
                "Calibration intercept (95% CI)": format_estimate_ci(
                    row, "Calibration_intercept"
                ),
                "Calibration slope (95% CI)": format_estimate_ci(
                    row, "Calibration_slope"
                ),
                "Threshold": round(float(row["Screening_threshold"]), 3),
                "Sensitivity (95% CI)": format_estimate_ci(row, "Sensitivity"),
                "Specificity (95% CI)": format_estimate_ci(row, "Specificity"),
                "PPV (95% CI)": format_estimate_ci(row, "PPV"),
                "NPV (95% CI)": format_estimate_ci(row, "NPV"),
            }
        )
    return pd.DataFrame(rows)


def sha256_file(path: Path, chunk_size: int = 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def make_threshold_table(
    results: dict[str, dict[str, Any]],
    target_sensitivity: float,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for task_key, result in results.items():
        spec: TaskSpec = result["spec"]
        for model_key, block in result["models"].items():
            info = block["threshold_info"]
            rows.append(
                {
                    "Analysis": spec.analysis_type,
                    "Task": task_key,
                    "Model": block["label"],
                    "Target_sensitivity": target_sensitivity,
                    **info,
                }
            )
    return pd.DataFrame(rows)


def create_readme(config: AnalysisConfig, paths: dict[str, Path]) -> None:
    text = f"""# Stroke landmark prediction pipeline output

## Run configuration
- Mode: {config.mode}
- Input: `{config.data_path}`
- Development period: {config.development_start} to the day before {config.temporal_cutoff}
- Temporal evaluation period: {config.temporal_cutoff} to the day before {config.study_end_exclusive}
- Landmarks: admission date +3 and +7 calendar days
- Primary tasks: Day 3 and Day 7 × no direct home discharge and LOS >21 days
- Models: L2-penalized logistic regression and LightGBM
- CV folds: {config.cv_folds}
- Optuna trials per LightGBM task: {config.n_optuna_trials}
- Bootstrap resamples: {config.n_bootstrap}
- Screening target sensitivity: {config.target_sensitivity:.2f}

## Important input assumption
The input workbook must contain one row per eligible hospitalization and landmark-specific
wide variables whose values were constructed upstream using only data available by each
landmark. This pipeline validates names, types, dates, outcomes, and landmark eligibility,
but it cannot retrospectively verify the source timestamps behind preconstructed day3_* and
day7_* columns.

## Output folders
- `models/`: locked fitted pipelines
- `predictions/`: development OOF and temporal-evaluation predictions
- `tables/`: CSV and Excel tables
- `figures/`: PDF and TIFF figures
- `diagnostics/`: cohort, outcome, missingness, and tuning audits
- `bootstrap/`: bootstrap distributions

## Manuscript alignment
The four primary tasks, chronological split, fold-specific preprocessing, target-sensitivity
thresholds, temporal evaluation, Day 7 operational matrix, paired landmark comparison,
LOS >28-day sensitivity analyses, and death-exclusion sensitivity analyses are all generated
from the same task definitions in this pipeline.
"""
    (paths["root"] / "README_output.md").write_text(text, encoding="utf-8")

# %% 
# =============================================================================
# 10. Three-stage orchestration
# =============================================================================

ANALYSIS_VERSION = "BMC_METHODS_ALIGNED_20260902"


def _clone_and_apply_mode(config: AnalysisConfig) -> AnalysisConfig:
    """Apply computational mode without mutating the caller's configuration."""
    from copy import deepcopy
    return deepcopy(config).apply_mode()


def setup_project_directories(config: AnalysisConfig) -> dict[str, dict[str, Path] | Path]:
    root = Path(config.output_dir)
    root.mkdir(parents=True, exist_ok=True)

    def stage_paths(stage_name: str) -> dict[str, Path]:
        stage_root = root / stage_name
        paths = {
            "root": stage_root,
            "models": stage_root / "models",
            "predictions": stage_root / "predictions",
            "tables": stage_root / "tables",
            "figures": stage_root / "figures",
            "diagnostics": stage_root / "diagnostics",
            "bootstrap": stage_root / "bootstrap",
            "bundles": stage_root / "bundles",
        }
        for path in paths.values():
            path.mkdir(parents=True, exist_ok=True)
        return paths

    return {
        "root": root,
        "main": stage_paths("01_main"),
        "sensitivity": stage_paths("02_sensitivity"),
        "publication": stage_paths("03_figures_tables"),
    }


def analysis_signature(config: AnalysisConfig) -> dict[str, Any]:
    """Settings that must remain identical across all three notebooks."""
    return {
        "analysis_version": ANALYSIS_VERSION,
        "mode": config.mode.upper(),
        "seed": config.seed,
        "id_col": config.id_col,
        "patient_id_col": config.patient_id_col,
        "admission_col": config.admission_col,
        "discharge_col": config.discharge_col,
        "los_col": config.los_col,
        "disposition_col": config.disposition_col,
        "development_start": config.development_start,
        "temporal_cutoff": config.temporal_cutoff,
        "study_end_exclusive": config.study_end_exclusive,
        "enforce_first_hospitalization": config.enforce_first_hospitalization,
        "known_duplicate_indices": list(config.known_duplicate_indices),
        "cv_folds": config.cv_folds,
        "n_optuna_trials": config.n_optuna_trials,
        "n_bootstrap": config.n_bootstrap,
        "cv_n_jobs": config.cv_n_jobs,
        "lgbm_n_jobs": config.lgbm_n_jobs,
        "target_sensitivity": config.target_sensitivity,
        "run_los28_day3": config.run_los28_day3,
        "run_los28_day7": config.run_los28_day7,
        "run_death_exclusion": config.run_death_exclusion,
        "lr_c_grid": list(config.lr_c_grid),
        "lgbm_search_space": {
            "learning_rate": [config.lgbm_learning_rate_min, config.lgbm_learning_rate_max],
            "n_estimators": [config.lgbm_estimators_min, config.lgbm_estimators_max],
            "max_depth": [config.lgbm_max_depth_min, config.lgbm_max_depth_max],
            "min_child_samples": [config.lgbm_min_child_samples_min, config.lgbm_min_child_samples_max],
            "subsample": [config.lgbm_subsample_min, config.lgbm_subsample_max],
            "colsample_bytree": [config.lgbm_colsample_min, config.lgbm_colsample_max],
            "regularization": [config.lgbm_regularization_min, config.lgbm_regularization_max],
        },
        "day3_continuous": list(DAY3_CONTINUOUS),
        "day3_binary": list(DAY3_BINARY),
        "day7_continuous": list(DAY7_CONTINUOUS),
        "day7_binary": list(DAY7_BINARY),
    }


def signature_hash(config: AnalysisConfig) -> str:
    payload = json.dumps(analysis_signature(config), ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def core_source_sha256() -> str:
    source_path = Path(__file__).resolve()
    return sha256_file(source_path)


def _stage_manifest_path(stage_paths: dict[str, Path]) -> Path:
    return stage_paths["root"] / "stage_manifest.json"


def write_stage_manifest(
    stage: str,
    config: AnalysisConfig,
    stage_paths: dict[str, Path],
    extra: Optional[dict[str, Any]] = None,
) -> Path:
    input_path = Path(config.data_path)
    manifest = {
        "analysis_version": ANALYSIS_VERSION,
        "stage": stage,
        "core_source_sha256": core_source_sha256(),
        "signature": analysis_signature(config),
        "signature_sha256": signature_hash(config),
        "input_file": str(input_path.resolve()),
        "input_file_size_bytes": input_path.stat().st_size,
        "input_file_sha256": sha256_file(input_path),
        "package_versions": package_versions(),
        "extra": extra or {},
    }
    path = _stage_manifest_path(stage_paths)
    path.write_text(json.dumps(json_safe(manifest), ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def validate_previous_stage(
    expected_stage: str,
    config: AnalysisConfig,
    stage_paths: dict[str, Path],
) -> dict[str, Any]:
    path = _stage_manifest_path(stage_paths)
    if not path.exists():
        raise FileNotFoundError(
            f"Required {expected_stage} manifest was not found: {path}. "
            "Run the notebooks in numerical order."
        )
    manifest = json.loads(path.read_text(encoding="utf-8"))
    if manifest.get("stage") != expected_stage:
        raise RuntimeError(f"Unexpected stage manifest in {path}")
    if manifest.get("analysis_version") != ANALYSIS_VERSION:
        raise RuntimeError("Analysis version differs between stages")
    if manifest.get("core_source_sha256") != core_source_sha256():
        raise RuntimeError(
            "The core analysis module changed after the preceding stage. "
            "Delete the output directory and rerun all three notebooks."
        )
    if manifest.get("signature_sha256") != signature_hash(config):
        raise RuntimeError(
            "Analysis settings differ from the preceding stage. Use identical CONFIG values "
            "in all three notebooks."
        )
    input_path = Path(config.data_path)
    if manifest.get("input_file_sha256") != sha256_file(input_path):
        raise RuntimeError("The input dataset changed after the preceding stage")
    return manifest


def _bundle_path(stage_paths: dict[str, Path], name: str) -> Path:
    return stage_paths["bundles"] / f"{name}.joblib"


def _save_bundle(obj: Any, stage_paths: dict[str, Path], name: str) -> Path:
    path = _bundle_path(stage_paths, name)
    joblib.dump(obj, path, compress=3)
    return path


def _load_bundle(stage_paths: dict[str, Path], name: str) -> Any:
    path = _bundle_path(stage_paths, name)
    if not path.exists():
        raise FileNotFoundError(f"Required analysis bundle was not found: {path}")
    return joblib.load(path)


def _task_set_qa(primary_results: dict[str, dict[str, Any]], config: AnalysisConfig) -> None:
    required = {"day3_nonhome", "day3_los21", "day7_nonhome", "day7_los21"}
    if set(primary_results) != required:
        raise AssertionError(f"Primary task set differs from Methods: {set(primary_results)}")
    for task_key, result in primary_results.items():
        if result["temporal"][config.id_col].duplicated().any():
            raise AssertionError(f"Duplicate temporal IDs in {task_key}")
    day7_nonhome_ids = set(primary_results["day7_nonhome"]["temporal"][config.id_col])
    day7_los_ids = set(primary_results["day7_los21"]["temporal"][config.id_col])
    if day7_nonhome_ids != day7_los_ids:
        raise AssertionError("Day 7 primary outcome cohorts differ")


def run_main_stage(config: AnalysisConfig) -> dict[str, Any]:
    """Notebook 01: fit the four primary models and lock all primary outputs."""
    config = _clone_and_apply_mode(config)
    np.random.seed(config.seed)
    project = setup_project_directories(config)
    paths = project["main"]
    logger = setup_logger(paths["root"] / "01_main_analysis.log")
    configure_matplotlib(config)

    logger.info("Stage 01 started | mode=%s", config.mode)
    data, audits = load_and_prepare_data(config, paths, logger)
    primary_specs, _ = build_task_specs(config)
    predictor_dictionary = make_predictor_dictionary()
    save_dataframe(predictor_dictionary, paths["tables"] / "predictor_dictionary.csv")

    primary_results: dict[str, dict[str, Any]] = {}
    for spec in primary_specs:
        primary_results[spec.key] = fit_prediction_task(data, spec, config, paths, logger)
    _task_set_qa(primary_results, config)

    performance = pd.DataFrame(
        [
            performance_row(result, model_key)
            for result in primary_results.values()
            for model_key in ["lr", "lgbm"]
        ]
    )
    publication_performance = make_publication_performance_table(performance)
    oof_performance = make_development_oof_performance(primary_results)
    thresholds = make_threshold_table(primary_results, config.target_sensitivity)
    hyperparameters = flatten_hyperparameters(primary_results)
    full_parameters = extract_full_model_parameters(primary_results)

    for name, table in {
        "primary_model_performance": performance,
        "primary_model_performance_publication_format": publication_performance,
        "development_oof_performance": oof_performance,
        "screening_thresholds": thresholds,
        "hyperparameters": hyperparameters,
        "full_fitted_model_parameters": full_parameters,
        "flow_counts_for_editable_Figure_1": audits["flow"],
    }.items():
        save_dataframe(table, paths["tables"] / f"{name}.csv")

    bundle = {
        "analysis_version": ANALYSIS_VERSION,
        "signature_sha256": signature_hash(config),
        "primary_results": primary_results,
        "performance": performance,
        "publication_performance": publication_performance,
        "oof_performance": oof_performance,
        "thresholds": thresholds,
        "hyperparameters": hyperparameters,
        "full_parameters": full_parameters,
    }
    bundle_path = _save_bundle(bundle, paths, "primary_results_bundle")
    manifest_path = write_stage_manifest(
        "main",
        config,
        paths,
        extra={
            "bundle": str(bundle_path),
            "N_total": len(data),
            "N_day3": int(data["flag_day3_landmark"].sum()),
            "N_day7": int(data["flag_day7_landmark"].sum()),
            "primary_tasks": [spec.key for spec in primary_specs],
        },
    )
    logger.info("Stage 01 completed | manifest=%s", manifest_path)
    return {"config": config, "paths": project, "data": data, "audits": audits, **bundle}


def run_sensitivity_stage(config: AnalysisConfig) -> dict[str, Any]:
    """Notebook 02: paired landmark comparison and prespecified sensitivity analysis."""
    config = _clone_and_apply_mode(config)
    np.random.seed(config.seed)
    project = setup_project_directories(config)
    main_paths = project["main"]
    paths = project["sensitivity"]
    logger = setup_logger(paths["root"] / "02_sensitivity_analysis.log")
    configure_matplotlib(config)

    validate_previous_stage("main", config, main_paths)
    main_bundle = _load_bundle(main_paths, "primary_results_bundle")
    primary_results = main_bundle["primary_results"]
    _task_set_qa(primary_results, config)

    logger.info("Stage 02 started | mode=%s", config.mode)
    data, audits = load_and_prepare_data(config, paths, logger)
    _, sensitivity_specs = build_task_specs(config)

    sensitivity_results: dict[str, dict[str, Any]] = {}
    for spec in sensitivity_specs:
        sensitivity_results[spec.key] = fit_prediction_task(data, spec, config, paths, logger)

    if sensitivity_results:
        sensitivity_performance = pd.DataFrame(
            [
                performance_row(result, model_key)
                for result in sensitivity_results.values()
                for model_key in ["lr", "lgbm"]
            ]
        )
        sensitivity_oof = make_development_oof_performance(sensitivity_results)
        sensitivity_thresholds = make_threshold_table(
            sensitivity_results, config.target_sensitivity
        )
        sensitivity_hyperparameters = flatten_hyperparameters(sensitivity_results)
    else:
        sensitivity_performance = pd.DataFrame()
        sensitivity_oof = pd.DataFrame()
        sensitivity_thresholds = pd.DataFrame()
        sensitivity_hyperparameters = pd.DataFrame()

    paired_summary, paired_bootstrap = paired_landmark_analysis(primary_results, config, paths)

    for name, table in {
        "sensitivity_model_performance": sensitivity_performance,
        "sensitivity_development_oof_performance": sensitivity_oof,
        "sensitivity_screening_thresholds": sensitivity_thresholds,
        "sensitivity_hyperparameters": sensitivity_hyperparameters,
        "paired_day3_day7_summary": paired_summary,
    }.items():
        if isinstance(table, pd.DataFrame):
            save_dataframe(table, paths["tables"] / f"{name}.csv")

    bundle = {
        "analysis_version": ANALYSIS_VERSION,
        "signature_sha256": signature_hash(config),
        "sensitivity_results": sensitivity_results,
        "sensitivity_performance": sensitivity_performance,
        "sensitivity_oof": sensitivity_oof,
        "sensitivity_thresholds": sensitivity_thresholds,
        "sensitivity_hyperparameters": sensitivity_hyperparameters,
        "paired_summary": paired_summary,
        "paired_bootstrap": paired_bootstrap,
    }
    bundle_path = _save_bundle(bundle, paths, "sensitivity_results_bundle")
    manifest_path = write_stage_manifest(
        "sensitivity",
        config,
        paths,
        extra={
            "bundle": str(bundle_path),
            "sensitivity_tasks": [spec.key for spec in sensitivity_specs],
            "paired_outcomes": ["nonhome", "los21"],
        },
    )
    logger.info("Stage 02 completed | manifest=%s", manifest_path)
    return {"config": config, "paths": project, "data": data, "audits": audits, **bundle}



def plot_dca_panels_common_scale(
    primary_results: dict[str, dict[str, Any]],
    config: AnalysisConfig,
) -> plt.Figure:
    thresholds = np.arange(
        config.dca_threshold_min,
        config.dca_threshold_max + config.dca_threshold_step / 2,
        config.dca_threshold_step,
    )

    curves: dict[str, dict[str, Any]] = {}
    finite_values: list[np.ndarray] = []

    for task_key, _ in panel_specs():
        r = primary_results[task_key]
        y = r["y_temp"].to_numpy()
        task_curves: dict[str, Any] = {}

        for model_key in ["lr", "lgbm"]:
            dca = decision_curve(
                y,
                r["models"][model_key]["temporal_prob"],
                thresholds,
            )
            task_curves[model_key] = dca
            finite_values.append(dca["model"].to_numpy())

        reference = decision_curve(
            y,
            r["models"]["lr"]["temporal_prob"],
            thresholds,
        )
        task_curves["reference"] = reference
        finite_values.append(reference["treat_all"].to_numpy())
        finite_values.append(reference["treat_none"].to_numpy())
        curves[task_key] = task_curves

    values = np.concatenate(finite_values)
    values = values[np.isfinite(values)]
    global_low = max(
        -0.10,
        min(-0.03, float(np.quantile(values, 0.02)) - 0.02),
    )
    global_high = min(
        0.70,
        max(0.05, float(np.quantile(values, 0.98)) + 0.02),
    )

    fig, axes = plt.subplots(
        2, 2,
        figsize=(mm_to_in(170), mm_to_in(160)),
    )

    for ax, (task_key, title) in zip(axes.ravel(), panel_specs()):
        r = primary_results[task_key]
        task_curves = curves[task_key]

        for model_key in ["lr", "lgbm"]:
            dca = task_curves[model_key]
            st = _MODEL_STYLE[model_key]
            ax.plot(
                dca["threshold"],
                dca["model"],
                color=st["color"],
                linestyle=st["linestyle"],
                linewidth=1.25,
            )

        reference = task_curves["reference"]
        ax.plot(
            reference["threshold"],
            reference["treat_all"],
            color=_REFERENCE_STYLE["treat_all"]["color"],
            linestyle=_REFERENCE_STYLE["treat_all"]["linestyle"],
            linewidth=1.0,
        )
        ax.plot(
            reference["threshold"],
            reference["treat_none"],
            color=_REFERENCE_STYLE["treat_none"]["color"],
            linestyle=_REFERENCE_STYLE["treat_none"]["linestyle"],
            linewidth=1.0,
        )

        ax.set_title(title, loc="left", fontweight="bold", pad=4)
        ax.set_xlabel("Threshold probability")
        ax.set_ylabel("Net benefit")
        ax.set_xlim(
            config.dca_threshold_min,
            config.dca_threshold_max,
        )
        ax.set_ylim(global_low, global_high)
        ax.set_xticks(
            np.arange(
                0.1,
                config.dca_threshold_max + 0.001,
                0.1,
            )
        )
        ax.tick_params(direction="out", length=2.5, width=0.6)

    # Single common legend for the whole figure.
    _common_model_legend(fig, include_references=True)
    fig.subplots_adjust(
        left=0.11, right=0.99, top=0.98, bottom=0.14,
        wspace=0.30, hspace=0.34,
    )
    return fig


def _draw_matrix_panel(
    ax: plt.Axes,
    summary: pd.DataFrame,
    panel_label: str,
    model_label: str,
) -> None:
    required_groups = {
        "Lower risk for both outcomes",
        "Higher non-home risk only",
        "Higher prolonged-hospitalization risk only",
        "Higher risk for both outcomes",
    }
    observed_groups = set(summary["matrix_group"].dropna())
    if observed_groups != required_groups:
        raise ValueError(
            f"Day 7 matrix summary does not contain the four prespecified strata: "
            f"{observed_groups}"
        )

    lookup = summary.set_index("matrix_group")
    threshold_nonhome = float(summary["threshold_nonhome"].dropna().iloc[0])
    threshold_los = float(summary["threshold_los21"].dropna().iloc[0])

    ax.set_xlim(0, 2)
    ax.set_ylim(0, 2)
    ax.set_aspect("equal")

    cells = [
        (
            "Lower risk for both outcomes",
            0, 0,
            "Lower risk for both",
        ),
        (
            "Higher prolonged-hospitalization risk only",
            1, 0,
            "Higher prolonged-LOS\nrisk only",
        ),
        (
            "Higher non-home risk only",
            0, 1,
            "Higher non-home\nrisk only",
        ),
        (
            "Higher risk for both outcomes",
            1, 1,
            "Higher risk for both",
        ),
    ]

    for group, x, y, short_title in cells:
        row = lookup.loc[group]

        rect = matplotlib.patches.Rectangle(
            (x, y),
            1,
            1,
            facecolor=_MATRIX_COLORS[group],
            edgecolor="#4D4D4D",
            linewidth=0.8,
        )
        ax.add_patch(rect)

        ax.text(
            x + 0.5,
            y + 0.76,
            short_title,
            ha="center",
            va="center",
            fontsize=8.1,
            fontweight="bold",
        )

        ax.text(
            x + 0.5,
            y + 0.39,
            (
                f"n = {int(row['n'])} ({100*row['cohort_percent']:.1f}%)\n"
                f"No-home discharge: {100*row['observed_nonhome_rate']:.1f}%\n"
                f"LOS >21 days: {100*row['observed_los21_rate']:.1f}%\n"
                f"Median LOS: {row['median_LOS']:.1f} days"
            ),
            ha="center",
            va="center",
            fontsize=6.9,
            linespacing=1.18,
        )

    ax.set_xticks([0.5, 1.5])
    ax.set_xticklabels(
        [
            f"Lower\n(p < {threshold_los:.3f})",
            f"Higher\n(p ≥ {threshold_los:.3f})",
        ]
    )
    ax.set_yticks([0.5, 1.5])
    ax.set_yticklabels(
        [
            f"Lower\n(p < {threshold_nonhome:.3f})",
            f"Higher\n(p ≥ {threshold_nonhome:.3f})",
        ]
    )

    ax.set_xlabel("Predicted risk of hospital stay >21 days")
    ax.set_ylabel("Predicted risk of no direct home discharge")

    ax.set_title(
        f"{panel_label} {model_label}",
        loc="left",
        fontweight="bold",
        pad=5,
    )

    for spine in ax.spines.values():
        spine.set_visible(False)



def plot_landmark_matrix_two_panel(
    summary_day3: pd.DataFrame,
    summary_day7: pd.DataFrame,
    *,
    model_label: str,
) -> plt.Figure:
    """
    Compare the same model at the two prespecified operational landmarks.

    (A) Day 3
    (B) Day 7
    """
    fig, axes = plt.subplots(
        1,
        2,
        figsize=(
            mm_to_in(170),
            mm_to_in(90),
        ),
    )

    _draw_matrix_panel(
        axes[0],
        summary_day3,
        "(A)",
        "Day 3",
    )

    _draw_matrix_panel(
        axes[1],
        summary_day7,
        "(B)",
        "Day 7",
    )

    # Avoid duplicate y-axis label on the right panel.
    axes[1].set_ylabel("")

    # Model identity is stated once for the full composite.
    fig.text(
        0.5,
        0.985,
        model_label,
        ha="center",
        va="top",
        fontsize=8.3,
        fontweight="bold",
    )

    fig.subplots_adjust(
        left=0.10,
        right=0.99,
        top=0.91,
        bottom=0.22,
        wspace=0.22,
    )

    return fig


def plot_day7_matrix_two_panel(
    summary_lr: pd.DataFrame,
    summary_lgbm: pd.DataFrame,
) -> plt.Figure:
    """
    Legacy wrapper retained only for backward compatibility.
    New publication code uses plot_landmark_matrix_two_panel().
    """
    fig, axes = plt.subplots(
        1,
        2,
        figsize=(
            mm_to_in(170),
            mm_to_in(90),
        ),
    )
    _draw_matrix_panel(
        axes[0],
        summary_lr,
        "(A)",
        "Penalized logistic regression",
    )
    _draw_matrix_panel(
        axes[1],
        summary_lgbm,
        "(B)",
        "LightGBM",
    )
    axes[1].set_ylabel("")
    fig.subplots_adjust(
        left=0.10,
        right=0.99,
        top=0.94,
        bottom=0.22,
        wspace=0.22,
    )
    return fig


def _make_figure_qa(figures_dir: Path) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for path in sorted(figures_dir.glob("*")):
        if path.suffix.lower() not in {".pdf", ".tiff", ".tif"}:
            continue
        size_bytes = path.stat().st_size
        rows.append(
            {
                "File": path.name,
                "Format": path.suffix.lower().lstrip("."),
                "Size_MB": size_bytes / 1024 / 1024,
                "Under_10_MB": size_bytes < 10 * 1024 * 1024,
            }
        )
    return pd.DataFrame(rows)


def _write_legends(paths: dict[str, Path]) -> None:
    text = """Figure 1. Landmark cohort construction and analysis workflow
Figure 1 is supplied separately as an editable Word figure. Numerical counts are exported as flow_counts_for_editable_Figure_1.csv.

Figure 2. Temporal-evaluation receiver operating characteristic curves
Receiver operating characteristic curves for no direct home discharge and hospital length of stay greater than 21 days at the Day 3 and Day 7 landmarks. AUROC values with 95% bootstrap confidence intervals are shown within each panel. The diagonal gray line indicates chance-level discrimination.

Figure 3. Calibration in the temporal-evaluation cohorts
Observed outcome proportions are plotted against mean predicted probabilities in approximately 10 quantile groups. The diagonal gray line indicates perfect calibration. Legends show model names only; calibration intercepts and slopes are reported separately within each panel.

Figure 4. Decision curve analysis in the temporal-evaluation cohorts
Net benefit is shown across threshold probabilities from 0.01 to 0.70 for penalized logistic regression, LightGBM, treat-all, and treat-none strategies. A common vertical scale and a single shared legend are used across panels.

Figure 5. Two-dimensional operational risk matrices at Day 3 and Day 7
(A) Day 3 and (B) Day 7 matrices from penalized logistic regression. Development-derived thresholds targeting approximately 80% sensitivity define four operational strata based on predicted risks of no direct home discharge and hospital stay greater than 21 days. Each cell reports cohort size, observed non-home discharge rate, observed prolonged-hospitalization rate, and median length of stay. The matrices characterize joint operational risk and are not validated treatment or disposition rules.

Additional Figure S1. Precision-recall curves in temporal evaluation
PR-AUC values with 95% bootstrap confidence intervals and outcome event rates are reported within each panel.

Additional Figure S2. Two-dimensional operational risk matrices using LightGBM
(A) Day 3 and (B) Day 7 matrices using the same development-derived screening-threshold procedure as the primary penalized logistic regression analysis.

Additional Figure S3. Penalized logistic regression coefficients
The 10 predictors with the largest absolute penalized coefficients are displayed for each primary prediction task.

Additional Figure S4. LightGBM SHAP summaries at Day 3
(A) No direct home discharge and (B) hospital stay >21 days. The 15 predictors with the largest mean absolute SHAP values are shown in each panel.

Additional Figure S5. LightGBM SHAP summaries at Day 7
(A) No direct home discharge and (B) hospital stay >21 days. The 15 predictors with the largest mean absolute SHAP values are shown in each panel. SHAP values describe model prediction contributions and should not be interpreted causally.
"""
    (paths["root"] / "figure_titles_and_legends.txt").write_text(text, encoding="utf-8")


def _write_project_readme(config: AnalysisConfig, project: dict[str, Any]) -> None:
    root = project["root"]
    text = f"""# Acute stroke Day 3/Day 7 landmark prediction pipeline

## Run order
1. `01_stroke_landmark_MAIN_BMC_METHODS_ALIGNED_20260902.ipynb`
2. `02_stroke_landmark_SENSITIVITY_BMC_METHODS_ALIGNED_20260902.ipynb`
3. `03_stroke_landmark_FIGURES_TABLES_BMC_METHODS_ALIGNED_20260902.ipynb`

All three notebooks must use exactly the same `CONFIG` values. Each stage verifies the input SHA-256 hash and analysis-signature hash from the preceding stage.

## Methods-aligned defaults
- Four primary tasks: Day 3/Day 7 × no direct home discharge/LOS >21 days
- Models: L2-penalized logistic regression and LightGBM
- Development: {config.development_start} to the day before {config.temporal_cutoff}
- Temporal evaluation: {config.temporal_cutoff} to the day before {config.study_end_exclusive}
- 5-fold stratified CV and binary log loss in FULL mode
- Development OOF thresholds targeting sensitivity closest to 80%
- 1,000 patient-level temporal bootstrap resamples in FULL mode
- Paired Day 3 versus Day 7 comparison among Day 7-eligible patients
- Day 7 LOS >28-day sensitivity analysis
- Day 3 and Day 7 two-dimensional matrices; penalized logistic regression is primary and LightGBM is supplementary
- No random holdout, no continuous LOS regression, no post-landmark predictors, no recalibration

## Disabled optional analyses
`run_los28_day3=False` and `run_death_exclusion=False` are the defaults because these analyses are not in the current Methods. Enabling either option requires corresponding Methods and Results reporting.

## Figure 1
The code does not generate Figure 1. It exports numerical cohort counts for the separately prepared editable Word flow diagram.

## Critical upstream assumptions
- `day3_*` and `day7_*` variables must have been constructed using only records available by the corresponding landmark.
- The exact discharge time is unavailable; discharge on the landmark date is treated as having reached that landmark.
- Missing `転帰` is mapped to direct home discharge according to the existing project data dictionary. Verify this coding before the final run.
- If `{config.patient_id_col}` is absent, the code records that first-eligible-hospitalization retention could not be independently verified and assumes it was completed upstream.

## Output folders
- `01_main`: locked primary models, predictions, bootstrap distributions, and QA
- `02_sensitivity`: paired comparison and sensitivity models
- `03_figures_tables`: publication figures, main tables, Additional file tables, legends, and final QA
"""
    (root / "README_output.md").write_text(text, encoding="utf-8")


def run_publication_stage(config: AnalysisConfig) -> dict[str, Any]:
    """Notebook 03: create publication tables and figures; performs no model fitting."""
    config = _clone_and_apply_mode(config)
    np.random.seed(config.seed)
    project = setup_project_directories(config)
    main_paths = project["main"]
    sensitivity_paths = project["sensitivity"]
    paths = project["publication"]
    logger = setup_logger(paths["root"] / "03_figures_tables.log")
    configure_matplotlib(config)

    validate_previous_stage("main", config, main_paths)
    validate_previous_stage("sensitivity", config, sensitivity_paths)
    main_bundle = _load_bundle(main_paths, "primary_results_bundle")
    sensitivity_bundle = _load_bundle(sensitivity_paths, "sensitivity_results_bundle")
    primary_results = main_bundle["primary_results"]
    sensitivity_results = sensitivity_bundle["sensitivity_results"]
    _task_set_qa(primary_results, config)

    logger.info("Stage 03 started | no model fitting")
    data, audits = load_and_prepare_data(config, paths, logger)
    predictor_dictionary = make_predictor_dictionary()

    primary_performance = main_bundle["performance"]
    primary_publication = main_bundle["publication_performance"]
    primary_oof = main_bundle["oof_performance"]
    primary_thresholds = main_bundle["thresholds"]
    paired_summary = sensitivity_bundle["paired_summary"]
    sensitivity_performance = sensitivity_bundle["sensitivity_performance"]

    matrix_day3_lr, matrix_summary_day3_lr = build_landmark_matrix(
        primary_results, "lr", config, 3
    )
    matrix_day7_lr, matrix_summary_day7_lr = build_landmark_matrix(
        primary_results, "lr", config, 7
    )
    matrix_day3_lgbm, matrix_summary_day3_lgbm = build_landmark_matrix(
        primary_results, "lgbm", config, 3
    )
    matrix_day7_lgbm, matrix_summary_day7_lgbm = build_landmark_matrix(
        primary_results, "lgbm", config, 7
    )

    save_dataframe(
        matrix_day3_lr,
        paths["predictions"] / "day3_operational_matrix_lr_patients.csv",
    )
    save_dataframe(
        matrix_day7_lr,
        paths["predictions"] / "day7_operational_matrix_lr_patients.csv",
    )
    save_dataframe(
        matrix_day3_lgbm,
        paths["predictions"] / "day3_operational_matrix_lgbm_patients.csv",
    )
    save_dataframe(
        matrix_day7_lgbm,
        paths["predictions"] / "day7_operational_matrix_lgbm_patients.csv",
    )

    save_dataframe(
        matrix_summary_day3_lr,
        paths["tables"] / "day3_operational_matrix_lr_summary.csv",
    )
    save_dataframe(
        matrix_summary_day7_lr,
        paths["tables"] / "day7_operational_matrix_lr_summary.csv",
    )
    save_dataframe(
        matrix_summary_day3_lgbm,
        paths["tables"] / "day3_operational_matrix_lgbm_summary.csv",
    )
    save_dataframe(
        matrix_summary_day7_lgbm,
        paths["tables"] / "day7_operational_matrix_lgbm_summary.csv",
    )

    characteristics_day3 = make_characteristics_table(
        data, 3, DAY3_CONTINUOUS, DAY3_BINARY, config
    )
    characteristics_day7 = make_characteristics_table(
        data, 7, DAY7_CONTINUOUS, DAY7_BINARY, config
    )
    missing_day3 = make_missingness_table(
        data, 3, list(dict.fromkeys(DAY3_CONTINUOUS + DAY3_BINARY))
    )
    missing_day7 = make_missingness_table(
        data, 7, list(dict.fromkeys(DAY7_CONTINUOUS + DAY7_BINARY))
    )

    all_results = {**primary_results, **sensitivity_results}
    hyperparameters = flatten_hyperparameters(all_results)
    full_model_parameters = extract_full_model_parameters(all_results)
    all_thresholds = make_threshold_table(all_results, config.target_sensitivity)
    lr_coefficients = pd.concat(
        [extract_lr_coefficients(result) for result in primary_results.values()],
        ignore_index=True,
    )
    lgbm_importance = pd.concat(
        [extract_lgbm_importance(result) for result in primary_results.values()],
        ignore_index=True,
    )

    shap_summary = pd.DataFrame()
    if config.run_shap:
        shap_tables: list[pd.DataFrame] = []
        for result in primary_results.values():
            shap_tables.append(calculate_shap_summary(result, config, paths))
        shap_summary = pd.concat(shap_tables, ignore_index=True)

    # No Figure 1 is generated here; only its counts are exported.
    save_dataframe(audits["flow"], paths["tables"] / "flow_counts_for_editable_Figure_1.csv")
    save_figure(plot_roc_panels(primary_results), paths["figures"] / "Figure_2_ROC", config.figure_dpi)
    save_figure(plot_calibration_panels(primary_results), paths["figures"] / "Figure_3_calibration", config.figure_dpi)
    save_figure(plot_dca_panels_common_scale(primary_results, config), paths["figures"] / "Figure_4_DCA", config.figure_dpi)
    save_figure(
        plot_landmark_matrix_two_panel(
            matrix_summary_day3_lr,
            matrix_summary_day7_lr,
            model_label="Penalized logistic regression",
        ),
        paths["figures"] / "Figure_5_operational_matrix_LR_Day3_Day7",
        config.figure_dpi,
    )
    save_figure(
        plot_pr_panels(primary_results),
        paths["figures"] / "Figure_S1_PR_curves",
        config.figure_dpi,
    )
    save_figure(
        plot_landmark_matrix_two_panel(
            matrix_summary_day3_lgbm,
            matrix_summary_day7_lgbm,
            model_label="LightGBM",
        ),
        paths["figures"] / "Figure_S2_operational_matrix_LightGBM_Day3_Day7",
        config.figure_dpi,
    )
    save_figure(
        plot_lr_coefficients(lr_coefficients),
        paths["figures"] / "Figure_S3_LR_coefficients",
        config.figure_dpi,
    )

    publication_tables = {
        "Flow_counts": audits["flow"],
        "Table1_Day3": characteristics_day3,
        "Table1_Day7": characteristics_day7,
        "Table2_Performance": primary_publication,
        "Table3_Day7_Matrix": matrix_summary_lr,
    }
    additional_tables = {
        "Predictor_dictionary": predictor_dictionary,
        "Outcome_audit": audits["outcome_audit"],
        "Disposition_audit": audits["disposition_audit"],
        "Missingness_Day3": missing_day3,
        "Missingness_Day7": missing_day7,
        "Development_OOF": primary_oof,
        "Thresholds": all_thresholds,
        "Hyperparameters": hyperparameters,
        "Full_model_parameters": full_model_parameters,
        "Paired_D3_D7": paired_summary,
        "Sensitivity": sensitivity_performance,
        "Matrix_LightGBM": matrix_summary_lgbm,
        "LR_coefficients": lr_coefficients,
        "LGBM_importance": lgbm_importance,
        "SHAP_summary": shap_summary,
        "Duplicate_audit": audits["duplicate_audit"],
        "First_hosp_audit": audits["first_hospitalization_audit"],
        "LOS_date_audit": audits["los_date_audit"],
        "Cumulative_landmark_QA": audits["cumulative_landmark_audit"],
    }
    main_workbook = paths["tables"] / "Main_tables_BMC.xlsx"
    additional_workbook = paths["tables"] / "Additional_file_tables_BMC.xlsx"
    write_excel_workbook(publication_tables, main_workbook)
    write_excel_workbook(additional_tables, additional_workbook)

    # Individual editable Excel files for manuscript preparation/submission handling.
    # No color fills or shading are applied to table cells.
    individual_main_tables = {
        "Table_1A_Day3_landmark_characteristics.xlsx": {"Table 1A": characteristics_day3},
        "Table_1B_Day7_landmark_characteristics.xlsx": {"Table 1B": characteristics_day7},
        "Table_2_model_performance.xlsx": {"Table 2": primary_publication},
        "Table_3_Day7_operational_matrix.xlsx": {"Table 3": matrix_summary_lr},
    }
    for filename, workbook_tables in individual_main_tables.items():
        write_excel_workbook(workbook_tables, paths["tables"] / filename)

    # Parallel CSV exports for transparent manuscript cross-checking.
    for name, table in {**publication_tables, **additional_tables}.items():
        if isinstance(table, pd.DataFrame) and (not table.empty or len(table.columns) > 0):
            save_dataframe(table, paths["tables"] / f"{name}.csv")

    _write_legends(paths)
    figure_qa = _make_figure_qa(paths["figures"])
    save_dataframe(figure_qa, paths["tables"] / "figure_file_QA.csv")
    if not figure_qa.empty and not figure_qa["Under_10_MB"].all():
        oversized = figure_qa.loc[~figure_qa["Under_10_MB"], "File"].tolist()
        raise RuntimeError(f"One or more figure files exceed 10 MB: {oversized}")

    manifest_path = write_stage_manifest(
        "publication",
        config,
        paths,
        extra={
            "main_workbook": str(main_workbook),
            "additional_workbook": str(additional_workbook),
            "figure_1_generated": False,
            "figure_files": figure_qa.to_dict(orient="records"),
        },
    )
    _write_project_readme(config, project)
    logger.info("Stage 03 completed | manifest=%s", manifest_path)
    return {
        "config": config,
        "paths": project,
        "data": data,
        "audits": audits,
        "primary_results": primary_results,
        "sensitivity_results": sensitivity_results,
        "primary_performance": primary_performance,
        "primary_publication": primary_publication,
        "paired_summary": paired_summary,
        "sensitivity_performance": sensitivity_performance,
        "matrix_summary_lr": matrix_summary_lr,
        "matrix_summary_lgbm": matrix_summary_lgbm,
        "figure_qa": figure_qa,
    }


def run_all_stages(config: AnalysisConfig) -> dict[str, Any]:
    run_main_stage(config)
    run_sensitivity_stage(config)
    return run_publication_stage(config)

# %% 
# =============================================================================
# Standalone notebook consistency signature
# =============================================================================
# All three standalone notebooks embed the same analysis engine.
# This fixed signature replaces the external core.py file hash used by the modular version.
STANDALONE_ANALYSIS_SOURCE_SHA256 = "6146868e35699f013ecaf9a8f4a1a1ef2c9d5c6ac4a1bf8c53dcafe48cf442bb"

def core_source_sha256() -> str:
    return STANDALONE_ANALYSIS_SOURCE_SHA256

print("Standalone analysis source SHA-256:", STANDALONE_ANALYSIS_SOURCE_SHA256)

# %% [markdown]
# ## User-editable analysis settings

# %% 

# =============================================================================
# 11. FAST publication stage
# =============================================================================
# This stage performs NO model fitting, NO Optuna optimization, and NO bootstrap.
# It reads the locked outputs from Notebook 01/02 and creates only manuscript
# figures/tables plus compact supplementary outputs.
#
# Heavy reproducibility objects (OOF predictions, full parameter dumps, audits)
# are exported as CSV rather than repeatedly styled in Excel.
# =============================================================================

import time


def _elapsed(start: float) -> str:
    return f"{time.perf_counter() - start:.1f} s"


def _log_step(label: str, start: float | None = None) -> float:
    if start is None:
        print(f"\n[START] {label}", flush=True)
        return time.perf_counter()
    print(f"[DONE ] {label} | {_elapsed(start)}", flush=True)
    return time.perf_counter()


def style_workbook_fast(path: Path, width_scan_rows: int = 200) -> None:
    """
    Lightweight BMC-compatible workbook styling.

    - no fill colors / no shading
    - bold header
    - wrapped text
    - Times New Roman for English/number cells, Meiryo for Japanese cells
    - column-width estimation uses at most the first width_scan_rows rows

    This is intentionally limited to compact manuscript/supplementary tables.
    """
    wb = openpyxl.load_workbook(path)
    align = Alignment(vertical="top", wrap_text=True)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        # Style cells once.
        for row_idx, row in enumerate(ws.iter_rows(), start=1):
            for cell in row:
                font_name = "Meiryo" if contains_japanese(cell.value) else "Times New Roman"
                cell.font = Font(name=font_name, bold=(row_idx == 1))
                cell.alignment = align

        # Estimate widths from header + a bounded number of rows.
        max_row_for_width = min(ws.max_row, max(2, width_scan_rows))
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(1, max_row_for_width + 1):
                value = ws.cell(row=row_idx, column=col_idx).value
                max_len = max(max_len, len("" if value is None else str(value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 45)

    wb.save(path)


def write_excel_workbook_fast(
    sheets: dict[str, pd.DataFrame],
    path: Path,
) -> None:
    """
    Write compact manuscript/supplementary tables to a single editable Excel file.
    """
    valid: dict[str, pd.DataFrame] = {}
    used: set[str] = set()

    for raw_name, df in sheets.items():
        if not isinstance(df, pd.DataFrame):
            continue
        if df.empty and len(df.columns) == 0:
            continue

        name = str(raw_name).replace("/", "_").replace("\\", "_")[:31] or "Sheet"
        base = name
        i = 1
        while name in used:
            suffix = f"_{i}"
            name = (base[:31 - len(suffix)] + suffix)[:31]
            i += 1
        used.add(name)
        valid[name] = df

    if not valid:
        raise ValueError(f"No nonempty tables to write: {path}")

    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, df in valid.items():
            df.to_excel(writer, sheet_name=name, index=False)

    style_workbook_fast(path)


def save_publication_figure(
    fig: plt.Figure,
    stem: Path,
    dpi: int = 300,
    export_tiff: bool = False,
) -> None:
    """
    Save vector PDF as the primary BMC submission format.
    TIFF is optional because BMC MIDM accepts PDF for vector figures.
    """
    stem.parent.mkdir(parents=True, exist_ok=True)

    fig.savefig(
        stem.with_suffix(".pdf"),
        bbox_inches="tight",
        pad_inches=0.02,
    )

    if export_tiff:
        fig.savefig(
            stem.with_suffix(".tiff"),
            dpi=dpi,
            bbox_inches="tight",
            pad_inches=0.02,
            pil_kwargs={"compression": "tiff_lzw"},
        )

    plt.close(fig)


def run_publication_stage_fast(
    config: AnalysisConfig,
    *,
    export_tiff: bool = False,
) -> dict[str, Any]:
    """
    FAST Notebook 03.

    What this function DOES:
    - validate Notebook 01/02 manifests
    - load locked model/sensitivity bundles
    - read the source dataset once to build Table 1 and missingness
    - create publication figures
    - create compact manuscript/supplementary Excel workbooks
    - export large reproducibility objects as CSV

    What this function DOES NOT do:
    - model fitting
    - Optuna
    - cross-validation
    - bootstrap confidence intervals
    - recalibration
    - SHAP (unless separately requested outside this fast stage)
    """
    overall_start = time.perf_counter()

    config = _clone_and_apply_mode(config)
    config.run_shap = False  # publication stage must remain lightweight/stable
    np.random.seed(config.seed)

    project = setup_project_directories(config)
    main_paths = project["main"]
    sensitivity_paths = project["sensitivity"]
    paths = project["publication"]
    logger = setup_logger(paths["root"] / "03_figures_tables_FAST.log")
    configure_matplotlib(config)

    # ------------------------------------------------------------------
    # 1. Validate and load already-locked analysis outputs
    # ------------------------------------------------------------------
    t = _log_step("Validate Notebook 01/02 outputs and load locked bundles")
    validate_previous_stage("main", config, main_paths)
    validate_previous_stage("sensitivity", config, sensitivity_paths)

    main_bundle = _load_bundle(main_paths, "primary_results_bundle")
    sensitivity_bundle = _load_bundle(sensitivity_paths, "sensitivity_results_bundle")

    primary_results = main_bundle["primary_results"]
    sensitivity_results = sensitivity_bundle["sensitivity_results"]
    _task_set_qa(primary_results, config)

    primary_performance = main_bundle["performance"]
    primary_publication = main_bundle["publication_performance"]
    primary_oof = main_bundle["oof_performance"]
    primary_thresholds = main_bundle["thresholds"]
    paired_summary = sensitivity_bundle["paired_summary"]
    sensitivity_performance = sensitivity_bundle["sensitivity_performance"]

    _log_step("Validate Notebook 01/02 outputs and load locked bundles", t)

    # ------------------------------------------------------------------
    # 2. Read source data ONCE for descriptive tables only
    # ------------------------------------------------------------------
    t = _log_step("Load source data once and build landmark descriptive tables")
    data, audits = load_and_prepare_data(config, paths, logger)

    predictor_dictionary = make_predictor_dictionary()
    characteristics_day3 = make_characteristics_table(
        data, 3, DAY3_CONTINUOUS, DAY3_BINARY, config
    )
    characteristics_day7 = make_characteristics_table(
        data, 7, DAY7_CONTINUOUS, DAY7_BINARY, config
    )
    missing_day3 = make_missingness_table(
        data, 3, list(dict.fromkeys(DAY3_CONTINUOUS + DAY3_BINARY))
    )
    missing_day7 = make_missingness_table(
        data, 7, list(dict.fromkeys(DAY7_CONTINUOUS + DAY7_BINARY))
    )
    _log_step("Load source data once and build landmark descriptive tables", t)

    # ------------------------------------------------------------------
    # 3. Derive compact reporting tables from already-fitted models
    # ------------------------------------------------------------------
    t = _log_step("Prepare locked model summaries and Day 3/Day 7 risk matrices")

    matrix_day3_lr, matrix_summary_day3_lr = build_landmark_matrix(
        primary_results,
        "lr",
        config,
        3,
    )
    matrix_day7_lr, matrix_summary_day7_lr = build_landmark_matrix(
        primary_results,
        "lr",
        config,
        7,
    )
    matrix_day3_lgbm, matrix_summary_day3_lgbm = build_landmark_matrix(
        primary_results,
        "lgbm",
        config,
        3,
    )
    matrix_day7_lgbm, matrix_summary_day7_lgbm = build_landmark_matrix(
        primary_results,
        "lgbm",
        config,
        7,
    )

    matrix_summary_lr = pd.concat(
        [
            matrix_summary_day3_lr,
            matrix_summary_day7_lr,
        ],
        ignore_index=True,
    )
    matrix_summary_lgbm = pd.concat(
        [
            matrix_summary_day3_lgbm,
            matrix_summary_day7_lgbm,
        ],
        ignore_index=True,
    )

    all_results = {**primary_results, **sensitivity_results}
    hyperparameters = flatten_hyperparameters(all_results)
    full_model_parameters = extract_full_model_parameters(all_results)
    all_thresholds = make_threshold_table(all_results, config.target_sensitivity)

    lr_coefficients = pd.concat(
        [extract_lr_coefficients(result) for result in primary_results.values()],
        ignore_index=True,
    )
    lgbm_importance = pd.concat(
        [extract_lgbm_importance(result) for result in primary_results.values()],
        ignore_index=True,
    )
    _log_step("Prepare locked model summaries and Day 3/Day 7 risk matrices", t)

    # ------------------------------------------------------------------
    # 4. Figures: PDF is primary; TIFF optional
    # ------------------------------------------------------------------
    t = _log_step("Generate publication figures")
    figure_jobs = [
        ("Figure_2_ROC", plot_roc_panels(primary_results)),
        ("Figure_3_calibration", plot_calibration_panels(primary_results)),
        ("Figure_4_DCA", plot_dca_panels_common_scale(primary_results, config)),
        (
            "Figure_5_operational_matrix_LR_Day3_Day7",
            plot_landmark_matrix_two_panel(
                matrix_summary_day3_lr,
                matrix_summary_day7_lr,
                model_label="Penalized logistic regression",
            ),
        ),
        ("Figure_S1_PR_curves", plot_pr_panels(primary_results)),
        (
            "Figure_S2_operational_matrix_LightGBM_Day3_Day7",
            plot_landmark_matrix_two_panel(
                matrix_summary_day3_lgbm,
                matrix_summary_day7_lgbm,
                model_label="LightGBM",
            ),
        ),
        ("Figure_S3_LR_coefficients", plot_lr_coefficients(lr_coefficients)),
    ]

    for stem_name, fig in figure_jobs:
        fstart = time.perf_counter()
        save_publication_figure(
            fig,
            paths["figures"] / stem_name,
            dpi=config.figure_dpi,
            export_tiff=export_tiff,
        )
        print(f"  saved {stem_name}.pdf | {_elapsed(fstart)}", flush=True)

    _log_step("Generate publication figures", t)

    # ------------------------------------------------------------------
    # 5. Compact manuscript Excel
    # ------------------------------------------------------------------
    t = _log_step("Write compact manuscript Excel tables")
    main_tables = {
        "Table1_Day3": characteristics_day3,
        "Table1_Day7": characteristics_day7,
        "Table2_Performance": primary_publication,
        "Table3_Operational_Matrix": matrix_summary_lr,
    }

    main_workbook = paths["tables"] / "Main_tables_BMC.xlsx"
    write_excel_workbook_fast(main_tables, main_workbook)
    _log_step("Write compact manuscript Excel tables", t)

    # ------------------------------------------------------------------
    # 6. Compact supplementary Excel
    # ------------------------------------------------------------------
    t = _log_step("Write compact supplementary Excel tables")
    supplementary_tables = {
        "Predictor_dictionary": predictor_dictionary,
        "Missingness_Day3": missing_day3,
        "Missingness_Day7": missing_day7,
        "Thresholds": all_thresholds,
        "Hyperparameters": hyperparameters,
        "Paired_D3_D7": paired_summary,
        "Sensitivity": sensitivity_performance,
        "Matrix_LightGBM_Day3_Day7": matrix_summary_lgbm,
        "LR_coefficients": lr_coefficients,
        "LGBM_importance": lgbm_importance,
    }

    additional_workbook = paths["tables"] / "Additional_file_tables_BMC.xlsx"
    write_excel_workbook_fast(supplementary_tables, additional_workbook)
    _log_step("Write compact supplementary Excel tables", t)

    # ------------------------------------------------------------------
    # 7. Reproducibility CSV only — deliberately NOT styled as Excel
    # ------------------------------------------------------------------
    t = _log_step("Export reproducibility/audit objects as CSV")
    reproducibility_dir = paths["tables"] / "reproducibility_csv"
    reproducibility_dir.mkdir(parents=True, exist_ok=True)

    reproducibility_tables = {
        "Flow_counts": audits["flow"],
        "Outcome_audit": audits["outcome_audit"],
        "Disposition_audit": audits["disposition_audit"],
        "Development_OOF": primary_oof,
        "Primary_thresholds": primary_thresholds,
        "Full_model_parameters": full_model_parameters,
        "Matrix_Day3_LR_patients": matrix_day3_lr,
        "Matrix_Day7_LR_patients": matrix_day7_lr,
        "Matrix_Day3_LightGBM_patients": matrix_day3_lgbm,
        "Matrix_Day7_LightGBM_patients": matrix_day7_lgbm,
        "Duplicate_audit": audits["duplicate_audit"],
        "First_hosp_audit": audits["first_hospitalization_audit"],
        "LOS_date_audit": audits["los_date_audit"],
        "Cumulative_landmark_QA": audits["cumulative_landmark_audit"],
    }

    for name, table in reproducibility_tables.items():
        if isinstance(table, pd.DataFrame) and (not table.empty or len(table.columns) > 0):
            save_dataframe(table, reproducibility_dir / f"{name}.csv")

    # Also export compact manuscript/supplementary tables as CSV for checking.
    for name, table in {**main_tables, **supplementary_tables}.items():
        if isinstance(table, pd.DataFrame) and (not table.empty or len(table.columns) > 0):
            save_dataframe(table, paths["tables"] / f"{name}.csv")

    _log_step("Export reproducibility/audit objects as CSV", t)

    # ------------------------------------------------------------------
    # 8. Legends and QA
    # ------------------------------------------------------------------
    t = _log_step("Write figure legends and run file QA")
    save_dataframe(audits["flow"], paths["tables"] / "flow_counts_for_editable_Figure_1.csv")
    _write_legends(paths)

    figure_qa = _make_figure_qa(paths["figures"])
    save_dataframe(figure_qa, paths["tables"] / "figure_file_QA.csv")

    if not figure_qa.empty and not figure_qa["Under_10_MB"].all():
        oversized = figure_qa.loc[~figure_qa["Under_10_MB"], "File"].tolist()
        raise RuntimeError(f"One or more figure files exceed 10 MB: {oversized}")

    _log_step("Write figure legends and run file QA", t)

    # ------------------------------------------------------------------
    # 9. Manifest
    # ------------------------------------------------------------------
    t = _log_step("Write publication manifest")
    manifest_path = write_stage_manifest(
        "publication",
        config,
        paths,
        extra={
            "fast_publication_stage": True,
            "model_fitting_performed": False,
            "optuna_performed": False,
            "bootstrap_performed": False,
            "shap_performed": False,
            "export_tiff": bool(export_tiff),
            "main_workbook": str(main_workbook),
            "additional_workbook": str(additional_workbook),
            "figure_1_generated": False,
            "figure_files": figure_qa.to_dict(orient="records"),
        },
    )
    _write_project_readme(config, project)
    _log_step("Write publication manifest", t)

    total = time.perf_counter() - overall_start
    print(f"\nFAST publication stage completed in {total/60:.2f} min", flush=True)
    print(f"Main tables: {main_workbook}", flush=True)
    print(f"Additional tables: {additional_workbook}", flush=True)
    print(f"Figures: {paths['figures']}", flush=True)

    return {
        "config": config,
        "paths": project,
        "data": data,
        "audits": audits,
        "primary_results": primary_results,
        "sensitivity_results": sensitivity_results,
        "primary_performance": primary_performance,
        "primary_publication": primary_publication,
        "paired_summary": paired_summary,
        "sensitivity_performance": sensitivity_performance,
        "matrix_summary_lr": matrix_summary_lr,
        "matrix_summary_lgbm": matrix_summary_lgbm,
        "matrix_summary_day3_lr": matrix_summary_day3_lr,
        "matrix_summary_day7_lr": matrix_summary_day7_lr,
        "matrix_summary_day3_lgbm": matrix_summary_day3_lgbm,
        "matrix_summary_day7_lgbm": matrix_summary_day7_lgbm,
        "figure_qa": figure_qa,
        "main_workbook": main_workbook,
        "additional_workbook": additional_workbook,
        "manifest_path": manifest_path,
    }

# %% 
# =============================================================================
# PORTABLE CONFIGURATION
# Set DATA_PATH and OUTPUT_DIR here, or adapt them to your local environment.
# =============================================================================

DATA_PATH = Path(
    r"data\Merge_20260901.xlsx"
)
OUTPUT_DIR = Path(
    r"outputs"
)

# Manuscript results: FULL
# Preliminary code check only: LIGHT
MODE = "FULL"

# Set to None only when the first eligible hospitalization per patient has already
# been fixed upstream and cannot be independently verified in this workbook.
PATIENT_ID_COL = "データ識別番号"

# Current Methods-aligned defaults.
# Native threads are capped at 1 for Windows/Jupyter stability.
RUN_LOS28_DAY3 = False
RUN_LOS28_DAY7 = True
RUN_DEATH_EXCLUSION = False
RUN_SHAP = False  # Safe default for Notebook 03; enable only for optional SHAP export

CONFIG = AnalysisConfig(
    data_path=DATA_PATH,
    output_dir=OUTPUT_DIR,
    mode=MODE,
    id_col="INDEX",
    patient_id_col=PATIENT_ID_COL,
    admission_col="入院日",
    discharge_col="退院日",
    los_col="在院日数",
    disposition_col="転帰",
    development_start="2018-04-01",
    temporal_cutoff="2023-04-01",
    study_end_exclusive="2025-04-01",
    enforce_first_hospitalization=True,
    known_duplicate_indices=(4768,),
    seed=20260902,
    cv_folds=5,
    n_optuna_trials=100,
    n_bootstrap=1000,
    cv_n_jobs=1,
    lgbm_n_jobs=1,
    target_sensitivity=0.80,
    dca_threshold_min=0.01,
    dca_threshold_max=0.70,
    dca_threshold_step=0.01,
    run_los28_day3=RUN_LOS28_DAY3,
    run_los28_day7=RUN_LOS28_DAY7,
    run_death_exclusion=RUN_DEATH_EXCLUSION,
    run_shap=RUN_SHAP,
    shap_max_patients=2000,
    figure_dpi=300,
    figure_font="Times New Roman",
)

if CONFIG.mode.upper() == "FULL":
    assert CONFIG.n_optuna_trials == 100, (
        f"FULL mode must use 100 Optuna trials; got {CONFIG.n_optuna_trials}."
    )

print(CONFIG)
print(f"Optuna trials per LightGBM task: {CONFIG.n_optuna_trials}")

# %% [markdown]
# ## Execute FAST publication stage
#
# This stage performs no model fitting, no Optuna, no bootstrap, and no SHAP. PDF is the primary BMC figure format. Set `EXPORT_TIFF = True` only when TIFF copies are required.

# %% 

# =============================================================================
# OPTIONAL FINAL STEP: LightGBM SHAP supplementary figures
# STANDARD SHAP -> INDIVIDUAL PANELS -> PIL COMBINATION
#
# No TIFF is created in this version.
#
# Workflow
# --------
# 1. Generate each panel independently with standard shap.plots.beeswarm().
# 2. Save each panel as:
#       - vector PDF
#       - high-resolution PNG (300 dpi) for combination
# 3. Combine the two high-resolution PNG panels with Pillow.
# 4. Save the final composite as:
#       - high-resolution PNG
#       - raster PDF
#
# Output
# ------
# Figure_S4_SHAP_Day3.png
# Figure_S4_SHAP_Day3.pdf
#
# Figure_S5_SHAP_Day7.png
# Figure_S5_SHAP_Day7.pdf
#
# Individual vector PDFs are retained in:
#   figures/individual_SHAP_panels/
#
# "Sum of N other features" is suppressed using:
#   group_remaining_features=False
#
# No model fitting, Optuna, bootstrap, or recalibration is performed.
# =============================================================================


def _prepare_standard_shap_task(
    task_key: str,
    primary_results: dict[str, dict[str, Any]],
    config: AnalysisConfig,
) -> dict[str, Any]:
    """Create a standard SHAP Explanation for one locked LightGBM model."""
    import shap

    task_result = primary_results[task_key]
    spec: TaskSpec = task_result["spec"]

    pipeline: Pipeline = task_result["models"]["lgbm"]["model"]
    preprocess: ColumnTransformer = pipeline.named_steps["preprocess"]
    model: LGBMClassifier = pipeline.named_steps["model"]

    x = task_result["x_temp"].copy()

    if len(x) > config.shap_max_patients:
        x = x.sample(
            n=config.shap_max_patients,
            random_state=stable_seed(
                config.seed,
                spec.key + "_shap_sample",
            ),
        )

    x_transformed = preprocess.transform(x)

    if hasattr(x_transformed, "toarray"):
        x_transformed = x_transformed.toarray()

    x_transformed = np.asarray(x_transformed, dtype=float)

    raw_names = [
        clean_feature_name(v)
        for v in preprocess.get_feature_names_out()
    ]
    display_names = [
        PUBLICATION_LABELS.get(name, name)
        for name in raw_names
    ]

    explainer = shap.TreeExplainer(model)
    raw_explanation = explainer(x_transformed)

    shap_values = np.asarray(raw_explanation.values, dtype=float)

    # Compatibility across SHAP/LightGBM versions.
    if shap_values.ndim == 3:
        shap_values = shap_values[:, :, -1]

    if shap_values.ndim != 2:
        raise ValueError(
            f"Unexpected SHAP dimensions for {task_key}: "
            f"{shap_values.shape}"
        )

    if shap_values.shape[1] != len(display_names):
        raise AssertionError(
            f"SHAP feature count mismatch for {task_key}: "
            f"{shap_values.shape[1]} vs {len(display_names)}"
        )

    explanation = shap.Explanation(
        values=shap_values,
        data=x_transformed,
        feature_names=display_names,
    )

    summary = pd.DataFrame(
        {
            "Task": spec.key,
            "Landmark": f"Day {spec.landmark_day}",
            "Outcome": spec.outcome_label,
            "Raw_variable": raw_names,
            "Variable": display_names,
            "Mean_absolute_SHAP": np.abs(shap_values).mean(axis=0),
            "Mean_SHAP": shap_values.mean(axis=0),
            "N_SHAP_patients": len(x),
        }
    ).sort_values(
        "Mean_absolute_SHAP",
        ascending=False,
    )

    return {
        "task_key": task_key,
        "spec": spec,
        "explanation": explanation,
        "summary": summary,
    }


def _save_standard_shap_panel(
    task_data: dict[str, Any],
    pdf_path: Path,
    png_path: Path,
    *,
    panel_label: str,
    panel_title: str,
    max_display: int = 15,
    dpi: int = 300,
) -> None:
    """
    Generate one standalone STANDARD SHAP beeswarm panel.

    - Standard shap.plots.beeswarm()
    - No custom jitter
    - No custom dot placement
    - No "Sum of N other features"
    """
    import shap

    pdf_path = Path(pdf_path)
    png_path = Path(png_path)

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    png_path.parent.mkdir(parents=True, exist_ok=True)

    fig, ax = plt.subplots(
        figsize=(
            mm_to_in(112),
            mm_to_in(125),
        )
    )

    shap.plots.beeswarm(
        task_data["explanation"],
        max_display=max_display,
        show=False,
        color_bar=True,
        ax=ax,
        plot_size=None,
        group_remaining_features=False,
    )

    ax.set_title(
        f"{panel_label} {panel_title}",
        loc="left",
        fontweight="bold",
        fontsize=9.0,
        pad=6,
    )

    ax.set_xlabel(
        "SHAP value (impact on model output)",
        fontsize=8.0,
    )

    ax.tick_params(
        axis="x",
        labelsize=7.0,
        direction="out",
        length=2.5,
        width=0.6,
    )
    ax.tick_params(
        axis="y",
        labelsize=7.0,
    )

    for extra_ax in fig.axes:
        if extra_ax is ax:
            continue
        extra_ax.tick_params(
            labelsize=7.0,
            length=0,
        )
        if extra_ax.get_ylabel():
            extra_ax.yaxis.label.set_size(7.5)

    fig.subplots_adjust(
        left=0.38,
        right=0.93,
        top=0.93,
        bottom=0.13,
    )

    fig.savefig(
        pdf_path,
        bbox_inches="tight",
        pad_inches=0.02,
    )

    fig.savefig(
        png_path,
        dpi=dpi,
        bbox_inches="tight",
        pad_inches=0.02,
        facecolor="white",
    )

    plt.close(fig)


def _combine_two_png_panels(
    left_png: Path,
    right_png: Path,
    output_stem: Path,
    *,
    target_width_mm: float = 170.0,
    gap_mm: float = 4.0,
    dpi: int = 300,
) -> dict[str, Path]:
    """
    Combine two high-resolution panel PNGs using Pillow.

    Final physical width = 170 mm at 300 dpi.
    Final outputs:
      - PNG
      - PDF
    """
    from PIL import Image

    left_png = Path(left_png)
    right_png = Path(right_png)
    output_stem = Path(output_stem)

    output_stem.parent.mkdir(parents=True, exist_ok=True)

    left = Image.open(left_png).convert("RGB")
    right = Image.open(right_png).convert("RGB")

    target_width_px = int(round(target_width_mm / 25.4 * dpi))
    gap_px = int(round(gap_mm / 25.4 * dpi))
    target_panel_width = int((target_width_px - gap_px) // 2)

    def resize_to_width(image: Image.Image, width: int) -> Image.Image:
        scale = width / image.width
        height = max(1, int(round(image.height * scale)))
        return image.resize(
            (width, height),
            resample=Image.Resampling.LANCZOS,
        )

    left_r = resize_to_width(left, target_panel_width)
    right_r = resize_to_width(right, target_panel_width)

    target_height = max(left_r.height, right_r.height)

    canvas = Image.new(
        "RGB",
        (target_width_px, target_height),
        color="white",
    )

    left_y = (target_height - left_r.height) // 2
    right_y = (target_height - right_r.height) // 2

    canvas.paste(left_r, (0, left_y))
    canvas.paste(right_r, (target_panel_width + gap_px, right_y))

    png_out = output_stem.with_suffix(".png")
    pdf_out = output_stem.with_suffix(".pdf")

    canvas.save(
        png_out,
        dpi=(dpi, dpi),
        optimize=True,
    )

    canvas.save(
        pdf_out,
        "PDF",
        resolution=float(dpi),
    )

    left.close()
    right.close()
    left_r.close()
    right_r.close()
    canvas.close()

    return {
        "png": png_out,
        "pdf": pdf_out,
    }


def export_lightgbm_shap_figures(
    primary_results: dict[str, dict[str, Any]],
    config: AnalysisConfig,
    figure_dir: Path,
    table_dir: Path,
    *,
    max_display: int = 15,
    panel_dpi: int = 300,
) -> pd.DataFrame:
    """
    Export standard SHAP beeswarms individually, then combine with Pillow.

    Final files:
      Figure_S4_SHAP_Day3.png
      Figure_S4_SHAP_Day3.pdf
      Figure_S5_SHAP_Day7.png
      Figure_S5_SHAP_Day7.pdf
    """
    figure_dir = Path(figure_dir)
    table_dir = Path(table_dir)

    figure_dir.mkdir(parents=True, exist_ok=True)
    table_dir.mkdir(parents=True, exist_ok=True)

    panel_dir = figure_dir / "individual_SHAP_panels"
    panel_dir.mkdir(parents=True, exist_ok=True)

    task_keys = [
        "day3_nonhome",
        "day3_los21",
        "day7_nonhome",
        "day7_los21",
    ]

    prepared: dict[str, dict[str, Any]] = {}
    summaries: list[pd.DataFrame] = []

    for task_key in task_keys:
        print(f"[SHAP] CALCULATE: {task_key}", flush=True)
        prepared[task_key] = _prepare_standard_shap_task(
            task_key,
            primary_results,
            config,
        )
        summaries.append(prepared[task_key]["summary"])

    individual_jobs = [
        (
            "day3_nonhome",
            "Figure_S4A_SHAP_Day3_nonhome",
            "(A)",
            "No direct home discharge",
        ),
        (
            "day3_los21",
            "Figure_S4B_SHAP_Day3_LOS21",
            "(B)",
            "Hospital stay >21 days",
        ),
        (
            "day7_nonhome",
            "Figure_S5A_SHAP_Day7_nonhome",
            "(A)",
            "No direct home discharge",
        ),
        (
            "day7_los21",
            "Figure_S5B_SHAP_Day7_LOS21",
            "(B)",
            "Hospital stay >21 days",
        ),
    ]

    panel_paths: dict[str, dict[str, Path]] = {}

    for task_key, stem, panel_label, panel_title in individual_jobs:
        pdf_path = panel_dir / f"{stem}.pdf"
        png_path = panel_dir / f"{stem}.png"

        print(f"[SHAP] PLOT: {stem}", flush=True)

        _save_standard_shap_panel(
            prepared[task_key],
            pdf_path,
            png_path,
            panel_label=panel_label,
            panel_title=panel_title,
            max_display=max_display,
            dpi=panel_dpi,
        )

        panel_paths[task_key] = {
            "pdf": pdf_path,
            "png": png_path,
        }

    print("[SHAP] COMBINE: Day 3", flush=True)
    day3_files = _combine_two_png_panels(
        panel_paths["day3_nonhome"]["png"],
        panel_paths["day3_los21"]["png"],
        figure_dir / "Figure_S4_SHAP_Day3",
        target_width_mm=170.0,
        gap_mm=4.0,
        dpi=panel_dpi,
    )

    print("[SHAP] COMBINE: Day 7", flush=True)
    day7_files = _combine_two_png_panels(
        panel_paths["day7_nonhome"]["png"],
        panel_paths["day7_los21"]["png"],
        figure_dir / "Figure_S5_SHAP_Day7",
        target_width_mm=170.0,
        gap_mm=4.0,
        dpi=panel_dpi,
    )

    print(f"[SHAP] SAVED Day 3 PNG: {day3_files['png'].name}", flush=True)
    print(f"[SHAP] SAVED Day 7 PNG: {day7_files['png'].name}", flush=True)
    print(f"[SHAP] SAVED Day 3 PDF: {day3_files['pdf'].name}", flush=True)
    print(f"[SHAP] SAVED Day 7 PDF: {day7_files['pdf'].name}", flush=True)

    for label, files in [("Day3", day3_files), ("Day7", day7_files)]:
        png_mb = files["png"].stat().st_size / (1024 ** 2)
        pdf_mb = files["pdf"].stat().st_size / (1024 ** 2)
        print(f"[SHAP] {label} PNG size: {png_mb:.2f} MB", flush=True)
        print(f"[SHAP] {label} PDF size: {pdf_mb:.2f} MB", flush=True)

    combined = pd.concat(summaries, ignore_index=True)

    combined.to_csv(
        table_dir / "SHAP_summary_all_tasks.csv",
        index=False,
        encoding="utf-8-sig",
    )

    return combined

# %% 
# PDF is sufficient for vector figures under BMC MIDM submission guidance.
# Set True only if you also want 300-dpi TIFF copies.
EXPORT_TIFF = False

with threadpool_limits(limits=1):
    PUBLICATION_RESULTS = run_publication_stage_fast(
        CONFIG,
        export_tiff=EXPORT_TIFF,
    )

# %% [markdown]
# ## LightGBM SHAP supplementary figures
#
# By default, SHAP values are not recalculated. The notebook searches previous analysis-output folders recursively for existing Day 3 and Day 7 SHAP PDF/PNG files and copies the newest matching files into the current publication directory as Figure S4 and Figure S5. If no prior SHAP figures exist, set `RECALCULATE_SHAP_IF_NOT_FOUND=True` to regenerate them.

# %% 

# =============================================================================
# SHAP supplementary figures
#
# Default behavior:
#   1. DO NOT recalculate SHAP.
#   2. Search previous analysis output folders recursively for existing
#      Day 3 / Day 7 SHAP PDF/PNG files.
#   3. Copy the newest matching files to the current publication folder using
#      the final numbering:
#          Figure S4 = SHAP Day 3
#          Figure S5 = SHAP Day 7
#   4. Only if nothing can be found, optionally recalculate by setting
#      RECALCULATE_SHAP_IF_NOT_FOUND = True.
# =============================================================================

from pathlib import Path
import shutil


RUN_SHAP_EXPORT = False
RECALCULATE_SHAP_IF_NOT_FOUND = False


if "PUBLICATION_RESULTS" not in globals():
    raise RuntimeError(
        "Run the FAST publication stage first so that "
        "PUBLICATION_RESULTS is available."
    )


publication_paths = PUBLICATION_RESULTS["paths"]["publication"]
figure_dir = Path(publication_paths["figures"])
table_dir = Path(publication_paths["tables"])

figure_dir.mkdir(
    parents=True,
    exist_ok=True,
)
table_dir.mkdir(
    parents=True,
    exist_ok=True,
)


# -----------------------------------------------------------------------------
# Search roots
# -----------------------------------------------------------------------------
#
# The previous SHAP figures were usually created under an earlier output
# directory. Searching only the current figure_dir is therefore insufficient.
#
# We search:
#   - current output root and its parent
#   - CONFIG.output_root if present
#   - parent folders around the current publication output
#
# The search is restricted to likely analysis-output roots to avoid scanning
# the entire drive.
# -----------------------------------------------------------------------------

candidate_roots = []


def _append_existing_root(path_like):
    if path_like is None:
        return
    try:
        p = Path(path_like).expanduser().resolve()
    except Exception:
        return

    if p.exists() and p.is_dir() and p not in candidate_roots:
        candidate_roots.append(p)


# Current publication hierarchy
_append_existing_root(figure_dir)
_append_existing_root(figure_dir.parent)
_append_existing_root(figure_dir.parent.parent)

if len(figure_dir.parents) >= 3:
    _append_existing_root(figure_dir.parents[2])

if len(figure_dir.parents) >= 4:
    _append_existing_root(figure_dir.parents[3])


# CONFIG output locations, if defined
for attr_name in [
    "output_root",
    "output_dir",
    "base_output",
    "project_root",
]:
    if hasattr(CONFIG, attr_name):
        value = getattr(CONFIG, attr_name)
        _append_existing_root(value)

        try:
            p = Path(value).expanduser().resolve()
            _append_existing_root(p.parent)
        except Exception:
            pass


print(
    "[SHAP] search roots:",
    flush=True,
)
for root in candidate_roots:
    print(
        f"  - {root}",
        flush=True,
    )


# -----------------------------------------------------------------------------
# Candidate filenames
# -----------------------------------------------------------------------------
#
# Include all numbering conventions used in earlier notebook versions.
# -----------------------------------------------------------------------------

DAY3_CANDIDATE_STEMS = [
    "Figure_S3_SHAP_Day3",
    "Figure_S4_SHAP_Day3",
]

DAY7_CANDIDATE_STEMS = [
    "Figure_S4_SHAP_Day7",
    "Figure_S5_SHAP_Day7",
]


def _find_existing_shap_file(
    stems,
    extension,
    *,
    exclude_destination=None,
):
    """
    Search recursively under candidate_roots and return the newest matching file.
    """
    matches = []

    for root in candidate_roots:
        for stem in stems:
            pattern = f"{stem}{extension}"

            try:
                for path in root.rglob(pattern):
                    try:
                        resolved = path.resolve()
                    except Exception:
                        resolved = path

                    if (
                        exclude_destination is not None
                        and resolved == exclude_destination.resolve()
                    ):
                        continue

                    if path.is_file():
                        matches.append(path)

            except (PermissionError, OSError):
                continue

    # Deduplicate paths.
    unique_matches = []
    seen = set()

    for path in matches:
        try:
            key = str(path.resolve())
        except Exception:
            key = str(path)

        if key not in seen:
            seen.add(key)
            unique_matches.append(path)

    if not unique_matches:
        return None

    # Prefer the most recently modified file.
    unique_matches.sort(
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )

    return unique_matches[0]


def _reuse_one_shap_figure(
    day_label,
    source_stems,
    destination_stem,
):
    """
    Reuse PDF and PNG independently.
    Returns True if at least a PDF was obtained.
    """
    found_any = False
    found_pdf = False

    for ext in [
        ".pdf",
        ".png",
    ]:
        destination = figure_dir / f"{destination_stem}{ext}"

        if destination.exists():
            print(
                f"[SHAP] already exists: {destination.name}",
                flush=True,
            )
            found_any = True

            if ext == ".pdf":
                found_pdf = True

            continue

        source = _find_existing_shap_file(
            source_stems,
            ext,
            exclude_destination=destination,
        )

        if source is None:
            print(
                f"[SHAP] no previous {day_label} {ext.upper()[1:]} found",
                flush=True,
            )
            continue

        shutil.copy2(
            source,
            destination,
        )

        print(
            f"[SHAP] reused {day_label}:",
            flush=True,
        )
        print(
            f"       source      = {source}",
            flush=True,
        )
        print(
            f"       destination = {destination}",
            flush=True,
        )

        found_any = True

        if ext == ".pdf":
            found_pdf = True

    return found_pdf, found_any


# IMPORTANT:
# Copy Day 7 first. An old Day 7 file may be named Figure_S4_SHAP_Day7,
# while the new Day 3 destination is Figure_S4_SHAP_Day3.
day7_pdf_ok, day7_any = _reuse_one_shap_figure(
    day_label="Day 7",
    source_stems=DAY7_CANDIDATE_STEMS,
    destination_stem="Figure_S5_SHAP_Day7",
)

day3_pdf_ok, day3_any = _reuse_one_shap_figure(
    day_label="Day 3",
    source_stems=DAY3_CANDIDATE_STEMS,
    destination_stem="Figure_S4_SHAP_Day3",
)


# -----------------------------------------------------------------------------
# Optional fallback recalculation
# -----------------------------------------------------------------------------

need_recalculation = not (
    day3_pdf_ok
    and day7_pdf_ok
)


if need_recalculation:
    print(
        "[SHAP] Existing PDF files were not found for both Day 3 and Day 7.",
        flush=True,
    )

    if RECALCULATE_SHAP_IF_NOT_FOUND:
        print(
            "[SHAP] Recalculating SHAP because "
            "RECALCULATE_SHAP_IF_NOT_FOUND=True.",
            flush=True,
        )

        primary_results_for_shap = PUBLICATION_RESULTS["primary_results"]

        with threadpool_limits(
            limits=1
        ):
            SHAP_SUMMARY = export_lightgbm_shap_figures(
                primary_results=primary_results_for_shap,
                config=CONFIG,
                figure_dir=figure_dir,
                table_dir=table_dir,
                max_display=15,
                panel_dpi=300,
            )

    else:
        print(
            "[SHAP] No recalculation performed.",
            flush=True,
        )
        print(
            "[SHAP] If these figures have never been generated in any previous "
            "output folder, set:",
            flush=True,
        )
        print(
            "       RECALCULATE_SHAP_IF_NOT_FOUND = True",
            flush=True,
        )

else:
    print(
        "[SHAP] Day 3 and Day 7 SHAP PDFs are ready. "
        "No SHAP recalculation was performed.",
        flush=True,
    )


# -----------------------------------------------------------------------------
# Final audit
# -----------------------------------------------------------------------------

expected_outputs = [
    figure_dir / "Figure_S4_SHAP_Day3.pdf",
    figure_dir / "Figure_S4_SHAP_Day3.png",
    figure_dir / "Figure_S5_SHAP_Day7.pdf",
    figure_dir / "Figure_S5_SHAP_Day7.png",
]

print(
    "\n[SHAP] final file audit",
    flush=True,
)

for path in expected_outputs:
    if path.exists():
        size_mb = path.stat().st_size / (1024 ** 2)
        print(
            f"  OK      {path.name}  ({size_mb:.2f} MB)",
            flush=True,
        )
    else:
        print(
            f"  MISSING {path.name}",
            flush=True,
        )
